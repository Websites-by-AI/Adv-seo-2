#!/usr/bin/env python3
"""Clinic Signal — dependency-free local server and safe public-site SEO auditor."""
from __future__ import annotations

import base64
import csv
import hashlib
import hmac
import ipaddress
import json
import math
import os
import re
import secrets
import smtplib
import socket

import requests
from bs4 import BeautifulSoup
import ssl
import time
from collections import defaultdict, deque
from email.message import EmailMessage
from html.parser import HTMLParser
from io import BytesIO, StringIO
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import parse_qs, quote_plus, unquote, urlencode, urljoin, urlparse
from urllib.request import HTTPRedirectHandler, Request, build_opener
from urllib.robotparser import RobotFileParser

ROOT = Path(__file__).resolve().parent
MAX_BODY = 2_000_000
USER_AGENT = "ClinicSignalAudit/1.1 (+public-business-seo-audit)"
SEND_ENABLED = os.getenv("SEND_ENABLED", "false").lower() == "true"
DRY_RUN = os.getenv("DRY_RUN", "true").lower() != "false"
SEND_LOG: deque[dict] = deque(maxlen=100)
RATE_BUCKETS: dict[str, deque[float]] = defaultdict(deque)
PDF_LINKS: dict[str, dict] = {}
PDF_LINK_TTL = max(300, min(int(os.getenv("PDF_LINK_TTL_SECONDS", "86400")), 604800))
PDF_LINK_LIMIT = max(10, min(int(os.getenv("PDF_LINK_LIMIT", "100")), 500))
ALLOWED_CHANNELS = {"whatsapp", "telegram", "bale", "rubika", "soroush", "eitaa", "email", "sms", "divar"}


def env_flag(name: str, default: bool = False) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "on"}


def integration_auth_error(headers, path: str) -> tuple[int, str] | None:
    """Validate Next.js → Python calls without exposing the shared token to a browser.

    X-Clinic-Signal-Internal makes verification mandatory for the Next.js gateway.
    CLINIC_SIGNAL_REQUIRE_AUTH=true optionally protects every /api route except signed PDF reads.
    The standalone Clinic Signal browser UI needs the latter left false unless a separate login
    or hosting-level access guard is configured.
    """
    if not path.startswith("/api/"):
        return None
    internal_call = str(headers.get("X-Clinic-Signal-Internal", "")).strip() == "1"
    globally_required = env_flag("CLINIC_SIGNAL_REQUIRE_AUTH", False)
    if not internal_call and not globally_required:
        return None
    if path in {"/api/shared-pdf"} and not internal_call:
        return None

    expected = os.getenv("CLINIC_SIGNAL_API_TOKEN", "").strip()
    if len(expected) < 24:
        return 503, "CLINIC_SIGNAL_API_TOKEN is missing or shorter than 24 characters."
    authorization = str(headers.get("Authorization", ""))
    provided = authorization[7:].strip() if authorization.startswith("Bearer ") else ""
    if not provided or not hmac.compare_digest(provided, expected):
        return 401, "Invalid Clinic Signal integration token."
    return None


try:
    from PIL import Image, ImageDraw, ImageFont, features as pil_features
    PILLOW_AVAILABLE = True
    RAQM_AVAILABLE = bool(pil_features.check("raqm"))
except Exception:
    PILLOW_AVAILABLE = False
    RAQM_AVAILABLE = False

try:
    import arabic_reshaper
    from bidi.algorithm import get_display as bidi_get_display
    BIDI_FALLBACK_AVAILABLE = True
except Exception:
    BIDI_FALLBACK_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False


def public_url(url: str) -> tuple[bool, str]:
    try:
        parsed = urlparse(url)
        if parsed.scheme not in {"http", "https"} or not parsed.hostname:
            return False, "Only public http/https URLs are allowed."
        if parsed.username or parsed.password:
            return False, "Credentials in URLs are not allowed."
        host = parsed.hostname.lower().rstrip(".")
        if host in {"localhost", "localhost.localdomain"} or host.endswith(".local"):
            return False, "Local hosts are blocked."
        infos = socket.getaddrinfo(host, parsed.port or (443 if parsed.scheme == "https" else 80))
        for info in infos:
            ip = ipaddress.ip_address(info[4][0])
            if not ip.is_global:
                return False, "Private, loopback and link-local destinations are blocked."
        return True, ""
    except Exception as exc:
        return False, f"Could not validate host: {type(exc).__name__}"


DIGIT_TRANSLATION = str.maketrans("۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩", "01234567890123456789")
EMAIL_PATTERN = re.compile(r"(?i)(?<![\w.+-])([a-z0-9][a-z0-9._%+-]{0,63}@[a-z0-9.-]+\.[a-z]{2,24})(?![\w.-])")
PHONE_CANDIDATE_PATTERN = re.compile(r"(?<!\d)(?:\+|00)?[\d۰-۹٠-٩][\d۰-۹٠-٩\s().-]{6,20}[\d۰-۹٠-٩](?!\d)")
CONTACT_PATH_PATTERN = re.compile(r"(?i)(contact|about|location|branch|clinic|office|support|تماس|درباره|شعب|آدرس)")
SOCIAL_HOSTS = ("instagram.com", "linkedin.com", "facebook.com", "youtube.com", "aparat.com", "t.me", "eitaa.com", "rubika.ir", "splus.ir", "bale.ai")
WHATSAPP_HOSTS = ("wa.me", "api.whatsapp.com", "web.whatsapp.com")


def normalize_public_phone(value: str) -> str:
    raw = str(value or "").translate(DIGIT_TRANSLATION).strip()
    digits = re.sub(r"\D", "", raw)
    if digits.startswith("0098"):
        digits = digits[2:]
    if digits.startswith("98") and 11 <= len(digits) <= 12:
        return "+" + digits
    if digits.startswith("0") and 10 <= len(digits) <= 11:
        return "+98" + digits[1:]
    if 8 <= len(digits) <= 15:
        return ("+" if raw.startswith("+") else "") + digits
    return ""


def extract_public_phones(text: str) -> list[str]:
    found = []
    seen = set()
    for match in PHONE_CANDIDATE_PATTERN.finditer(str(text or "")):
        phone = normalize_public_phone(match.group(0))
        digits = re.sub(r"\D", "", phone)
        # Reject obvious dates and repeated placeholders while retaining public landlines/mobiles.
        if not phone or len(set(digits)) < 3 or digits.startswith(("139", "140")) and len(digits) <= 8:
            continue
        if phone not in seen:
            seen.add(phone)
            found.append(phone)
    return found[:20]


def extract_public_contact_signals(html: str, base_url: str) -> dict:
    """Extract public business contact signals only; never submit forms or access private pages."""
    soup = BeautifulSoup(html or "", "html.parser")
    phones, emails, whatsapp_links, social_links, contact_pages, tags = set(), set(), set(), set(), set(), set()
    base = urlparse(base_url)
    base_host = (base.hostname or "").lower().removeprefix("www.")

    for anchor in soup.find_all("a", href=True):
        raw_href = str(anchor.get("href", "")).strip()
        lower = raw_href.lower()
        if lower.startswith("tel:"):
            phone = normalize_public_phone(raw_href[4:].split("?", 1)[0])
            if phone:
                phones.add(phone)
            continue
        if lower.startswith("mailto:"):
            email = raw_href[7:].split("?", 1)[0].strip().lower()
            if EMAIL_PATTERN.fullmatch(email):
                emails.add(email)
            continue
        absolute = urljoin(base_url, raw_href)
        parsed = urlparse(absolute)
        host = (parsed.hostname or "").lower().removeprefix("www.")
        if parsed.scheme not in {"http", "https"} or not host:
            continue
        if host in WHATSAPP_HOSTS or host.endswith(".whatsapp.com"):
            whatsapp_links.add(absolute.split("#", 1)[0])
            query = parse_qs(parsed.query)
            candidate = parsed.path.strip("/").split("/", 1)[0] if host == "wa.me" else (query.get("phone") or [""])[0]
            phone = normalize_public_phone(candidate)
            if phone:
                phones.add(phone)
            continue
        if any(host == item or host.endswith("." + item) for item in SOCIAL_HOSTS):
            social_links.add(absolute.split("#", 1)[0])
        if host == base_host and CONTACT_PATH_PATTERN.search(unquote(parsed.path)):
            contact_pages.add(absolute.split("#", 1)[0])

    visible_text = soup.get_text(" ", strip=True)
    phones.update(extract_public_phones(visible_text))
    emails.update(x.group(1).lower() for x in EMAIL_PATTERN.finditer(visible_text))

    for meta in soup.find_all("meta"):
        key = str(meta.get("name") or meta.get("property") or "").lower()
        value = str(meta.get("content", "")).strip()
        if key in {"keywords", "news_keywords", "article:tag"}:
            tags.update(x.strip()[:80] for x in re.split(r"[,،|]", value) if x.strip())

    addresses = []
    address_tag = soup.find("address")
    if address_tag:
        value = " ".join(address_tag.get_text(" ", strip=True).split())
        if value:
            addresses.append(value[:500])

    def walk_json(value):
        if isinstance(value, list):
            for item in value:
                walk_json(item)
        elif isinstance(value, dict):
            telephone = value.get("telephone")
            if telephone:
                phone = normalize_public_phone(str(telephone))
                if phone:
                    phones.add(phone)
            email = str(value.get("email", "")).removeprefix("mailto:").strip().lower()
            if email and EMAIL_PATTERN.fullmatch(email):
                emails.add(email)
            for key in ("keywords", "medicalSpecialty", "serviceType", "knowsAbout"):
                values = value.get(key, [])
                if not isinstance(values, list):
                    values = re.split(r"[,،|]", str(values))
                tags.update(str(x).strip()[:80] for x in values if str(x).strip())
            address = value.get("address")
            if isinstance(address, dict):
                rendered = "، ".join(str(address.get(k, "")).strip() for k in ("addressCountry", "addressRegion", "addressLocality", "streetAddress", "postalCode") if address.get(k))
                if rendered:
                    addresses.append(rendered[:500])
            for child in value.values():
                if isinstance(child, (dict, list)):
                    walk_json(child)

    for script in soup.find_all("script", attrs={"type": "application/ld+json"}):
        try:
            walk_json(json.loads(script.get_text(strip=True) or "{}"))
        except Exception:
            continue

    whatsapp_number = ""
    for link in sorted(whatsapp_links):
        parsed = urlparse(link)
        query = parse_qs(parsed.query)
        candidate = parsed.path.strip("/").split("/", 1)[0] if (parsed.hostname or "").lower() == "wa.me" else (query.get("phone") or [""])[0]
        whatsapp_number = normalize_public_phone(candidate)
        if whatsapp_number:
            break

    return {
        "phoneNumbers": sorted(phones)[:20],
        "emails": sorted(emails)[:20],
        "whatsappLinks": sorted(whatsapp_links)[:10],
        "whatsappNumber": whatsapp_number,
        "socialLinks": sorted(social_links)[:20],
        "contactPageCandidates": sorted(contact_pages)[:10],
        "tags": sorted(tags, key=str.casefold)[:40],
        "addresses": list(dict.fromkeys(addresses))[:8],
    }


def enrich_public_business_contacts(url: str, max_pages: int = 3) -> dict:
    """Crawl a few same-origin public contact/about pages, respecting robots.txt."""
    if not urlparse(url).scheme:
        url = "https://" + url.strip()
    ok, reason = public_url(url)
    if not ok:
        raise ValueError(reason)
    max_pages = max(1, min(int(max_pages or 3), 5))
    status, final_url, html, content_type, elapsed = fetch(url, timeout=15, limit=1_500_000)
    if status >= 400 or ("html" not in content_type.lower() and "<html" not in html[:1000].lower()):
        raise ValueError(f"Website returned HTTP {status} or non-HTML content.")

    base_host = (urlparse(final_url).hostname or "").lower().removeprefix("www.")
    aggregate = extract_public_contact_signals(html, final_url)
    pages = [{"url": final_url, "status": status, "title": ""}]
    parser = AuditParser(final_url)
    parser.feed(html)
    pages[0]["title"] = parser.title

    candidates = list(aggregate.pop("contactPageCandidates", []))
    for page_url in candidates:
        if len(pages) >= max_pages:
            break
        host = (urlparse(page_url).hostname or "").lower().removeprefix("www.")
        if host != base_host or not robots_allows(page_url):
            continue
        try:
            page_status, page_final, page_html, page_type, _ = fetch(page_url, timeout=12, limit=1_000_000)
            if page_status >= 400 or ("html" not in page_type.lower() and "<html" not in page_html[:1000].lower()):
                continue
            signals = extract_public_contact_signals(page_html, page_final)
            for key in ("phoneNumbers", "emails", "whatsappLinks", "socialLinks", "tags", "addresses"):
                aggregate[key] = list(dict.fromkeys([*aggregate.get(key, []), *signals.get(key, [])]))[:40]
            if not aggregate.get("whatsappNumber") and signals.get("whatsappNumber"):
                aggregate["whatsappNumber"] = signals["whatsappNumber"]
            page_parser = AuditParser(page_final)
            page_parser.feed(page_html)
            pages.append({"url": page_final, "status": page_status, "title": page_parser.title})
        except Exception:
            continue

    phones = aggregate.get("phoneNumbers", [])
    emails = aggregate.get("emails", [])
    return {
        "ok": True,
        "requestedUrl": url,
        "finalUrl": final_url,
        "status": status,
        "elapsedSeconds": elapsed,
        "primaryPhone": phones[0] if phones else "",
        "primaryEmail": emails[0] if emails else "",
        **aggregate,
        "pagesChecked": pages,
        "disclaimer": "Public business contact signals only. Verify ownership and recipient consent before outreach; no forms, accounts or patient data were accessed.",
    }


class SafeRedirect(HTTPRedirectHandler):
    def redirect_request(self, req, fp, code, msg, headers, newurl):
        absolute = urljoin(req.full_url, newurl)
        ok, reason = public_url(absolute)
        if not ok:
            raise URLError(f"Unsafe redirect blocked: {reason}")
        return super().redirect_request(req, fp, code, msg, headers, absolute)


class AuditParser(HTMLParser):
    def __init__(self, base_url: str):
        super().__init__(convert_charrefs=True)
        self.base_url = base_url
        self.title = ""
        self.description = ""
        self.h1: list[str] = []
        self.canonical = ""
        self.lang = ""
        self.viewport = False
        self.og_title = False
        self.schema_blocks = 0
        self.schema_types: set[str] = set()
        self.links: set[str] = set()
        self.social_links: set[str] = set()
        self.phone_links: set[str] = set()
        self.email_links: set[str] = set()
        self.text_chars = 0
        self.text_words = 0
        self.phone_signal = False
        self.address_signal = False
        self.map_or_social_signal = False
        self._in_title = False
        self._in_h1 = False
        self._in_schema = False
        self._buf: list[str] = []

    @staticmethod
    def attrs_dict(attrs):
        return {str(k).lower(): (v or "") for k, v in attrs}

    def handle_starttag(self, tag, attrs):
        tag = tag.lower()
        a = self.attrs_dict(attrs)
        if tag == "html":
            self.lang = a.get("lang", "")
        elif tag == "title":
            self._in_title = True
            self._buf = []
        elif tag == "h1":
            self._in_h1 = True
            self._buf = []
        elif tag == "meta":
            name = a.get("name", "").lower()
            prop = a.get("property", "").lower()
            if name == "description" and not self.description:
                self.description = a.get("content", "").strip()
            if name == "viewport":
                self.viewport = True
            if prop == "og:title":
                self.og_title = True
        elif tag == "link" and "canonical" in a.get("rel", "").lower():
            self.canonical = a.get("href", "").strip()
        elif tag == "script" and "ld+json" in a.get("type", "").lower():
            self._in_schema = True
            self._buf = []
            self.schema_blocks += 1
        elif tag == "a" and a.get("href"):
            raw_href = a["href"].strip()
            raw = raw_href.lower()
            if raw.startswith("tel:"):
                self.phone_links.add(raw_href[4:].strip())
                self.phone_signal = True
                return
            if raw.startswith("mailto:"):
                self.email_links.add(raw_href[7:].split("?", 1)[0].strip())
                return
            href = urljoin(self.base_url, raw_href)
            parsed = urlparse(href)
            if parsed.scheme in {"http", "https"}:
                clean_href = href.split("#", 1)[0]
                self.links.add(clean_href)
                if any(x in parsed.netloc.lower() for x in ("instagram.com", "linkedin.com", "facebook.com", "youtube.com", "aparat.com", "t.me", "wa.me", "eitaa.com", "rubika.ir", "splus.ir", "bale.ai")):
                    self.social_links.add(clean_href)
            if any(x in raw for x in ("instagram.com", "maps.google", "goo.gl/maps", "wa.me", "t.me", "eitaa.com", "rubika.ir", "splus.ir", "bale.ai")):
                self.map_or_social_signal = True

    def handle_endtag(self, tag):
        tag = tag.lower()
        text = " ".join("".join(self._buf).split())
        if tag == "title" and self._in_title:
            self.title = text
            self._in_title = False
        elif tag == "h1" and self._in_h1:
            if text:
                self.h1.append(text)
            self._in_h1 = False
        elif tag == "script" and self._in_schema:
            for typ in re.findall(r'"@type"\s*:\s*"([^"]+)"', text):
                self.schema_types.add(typ)
            self._in_schema = False
        self._buf = []

    def handle_data(self, data):
        clean = " ".join(data.split())
        if clean:
            self.text_chars += len(clean)
            self.text_words += len(clean.split())
            low = clean.lower()
            if re.search(r'(?:\+?98|0)?21[-\s]?\d{5,8}', clean) or re.search(r'09\d{9}', clean):
                self.phone_signal = True
            if any(x in low for x in ("تهران", "آدرس", "address", "خیابان", "street")):
                self.address_signal = True
        if self._in_title or self._in_h1 or self._in_schema:
            self._buf.append(data)


def fetch(url: str, timeout: int = 14, limit: int = MAX_BODY):
    ok, reason = public_url(url)
    if not ok:
        raise ValueError(reason)
    opener = build_opener(SafeRedirect())
    req = Request(url, headers={"User-Agent": USER_AGENT, "Accept": "text/html,application/xhtml+xml"})
    started = time.monotonic()
    try:
        with opener.open(req, timeout=timeout) as response:
            final_url = response.geturl()
            ok, reason = public_url(final_url)
            if not ok:
                raise ValueError(f"Unsafe final URL: {reason}")
            body = response.read(limit + 1)
            if len(body) > limit:
                body = body[:limit]
            content_type = response.headers.get("Content-Type", "")
            charset = response.headers.get_content_charset() or "utf-8"
            text = body.decode(charset, errors="replace")
            return int(response.status), final_url, text, content_type, round(time.monotonic() - started, 2)
    except HTTPError as exc:
        try:
            body = exc.read(min(limit, 200_000)).decode("utf-8", errors="replace")
        except Exception:
            body = ""
        return int(exc.code), exc.geturl(), body, exc.headers.get("Content-Type", ""), round(time.monotonic() - started, 2)


def endpoint_exists(url: str, endpoint: str) -> bool:
    try:
        status, _, text, content_type, _ = fetch(urljoin(url, endpoint), timeout=7, limit=400_000)
        if status != 200:
            return False
        head = text[:2000].lower()
        if endpoint == "robots.txt":
            return "user-agent" in head or "sitemap:" in head
        return "<urlset" in head or "<sitemapindex" in head
    except Exception:
        return False


def score_audit(status: int, final_url: str, p: AuditParser, robots: bool, sitemap: bool):
    score = 0
    issues: list[str] = []
    wins: list[str] = []

    if status == 200:
        score += 25
        wins.append("Homepage is reachable with HTTP 200")
    else:
        issues.append(f"Homepage returned HTTP {status}")

    if 20 <= len(p.title) <= 65:
        score += 5
    else:
        issues.append("Title is missing or outside the useful 20–65 character range")
    if 70 <= len(p.description) <= 170:
        score += 5
    else:
        issues.append("Meta description is missing or outside the useful range")
    if len(p.h1) == 1:
        score += 5
    else:
        issues.append(f"Expected one H1; found {len(p.h1)}")
    if p.canonical:
        score += 4
    else:
        issues.append("Canonical tag was not found")
    if p.viewport:
        score += 3
    else:
        issues.append("Mobile viewport tag was not found")
    if p.schema_blocks:
        score += 4
        wins.append("Structured data was detected")
    else:
        issues.append("No JSON-LD structured data was detected")
    if p.og_title:
        score += 2

    if p.text_chars >= 1800:
        score += 10
    elif p.text_chars >= 700:
        score += 6
    else:
        issues.append("Homepage has little crawlable text")
    if len(p.links) >= 20:
        score += 10
    elif len(p.links) >= 8:
        score += 6
    else:
        issues.append("Internal linking appears thin")

    if robots:
        score += 6
    else:
        issues.append("A valid robots.txt was not confirmed")
    if sitemap:
        score += 6
    else:
        issues.append("A valid sitemap.xml was not confirmed")
    if final_url.startswith("https://"):
        score += 3
    else:
        issues.append("Final page is not HTTPS")

    medical = any(x.lower() in {"medicalclinic", "medicalbusiness", "physician", "dermatology"} for x in p.schema_types)
    if p.phone_signal:
        score += 5
    else:
        issues.append("Public phone signal was not detected on the homepage")
    if p.address_signal:
        score += 5
    else:
        issues.append("Address/location signal was not detected")
    if medical or p.map_or_social_signal:
        score += 5
    else:
        issues.append("Medical entity or map/social identity signal is weak")

    return min(100, score), issues[:8], wins[:5]


def audit(url: str):
    if not urlparse(url).scheme:
        url = "https://" + url.strip()
    started = time.monotonic()
    try:
        status, final_url, html, content_type, elapsed = fetch(url)
    except (URLError, ssl.SSLCertVerificationError) as exc:
        message = str(exc)
        if "CERTIFICATE_VERIFY_FAILED" not in message and "certificate" not in message.lower() and "ssl" not in message.lower():
            raise
        return {"ok": True, "requestedUrl": url, "status": 0, "finalUrl": url,
                "elapsedSeconds": round(time.monotonic()-started, 2), "totalSeconds": round(time.monotonic()-started, 2),
                "title": "", "titleLength": 0, "description": "", "descriptionLength": 0,
                "h1Count": 0, "h1": [], "canonical": "", "lang": "", "viewport": False,
                "schemaBlocks": 0, "schemaTypes": [], "wordCount": 0,
                "internalLinks": 0, "externalLinks": 0, "internalLinkSamples": [], "externalLinkSamples": [],
                "socialLinks": [], "phoneLinks": [], "emailLinks": [], "whatsappLinks": [],
                "whatsappNumber": "", "tags": [], "publicAddresses": [], "contactPageCandidates": [], "textCharacters": 0,
                "robots": False, "sitemap": False, "seoScore": 5, "sslError": True,
                "issues": ["SSL certificate validation failed or the certificate has expired", message[:300]],
                "wins": [], "checkedAt": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
                "disclaimer": "SSL failure is a measured critical availability issue. Confirm from the target market before outreach."}
    if "html" not in content_type.lower() and "<html" not in html[:1000].lower():
        raise ValueError("The URL did not return an HTML page.")
    parser = AuditParser(final_url)
    parser.feed(html)
    contacts = extract_public_contact_signals(html, final_url)
    robots = endpoint_exists(final_url, "robots.txt")
    sitemap = endpoint_exists(final_url, "sitemap.xml")
    score, issues, wins = score_audit(status, final_url, parser, robots, sitemap)
    base_host = (urlparse(final_url).hostname or "").lower().removeprefix("www.")
    internal_urls, external_urls = [], []
    for link in sorted(parser.links):
        host = (urlparse(link).hostname or "").lower().removeprefix("www.")
        if host == base_host:
            internal_urls.append(link)
        else:
            external_urls.append(link)
    return {
        "ok": True,
        "requestedUrl": url,
        "status": status,
        "finalUrl": final_url,
        "elapsedSeconds": elapsed,
        "totalSeconds": round(time.monotonic() - started, 2),
        "title": parser.title,
        "titleLength": len(parser.title),
        "description": parser.description,
        "descriptionLength": len(parser.description),
        "h1Count": len(parser.h1),
        "h1": parser.h1[:3],
        "canonical": parser.canonical,
        "lang": parser.lang,
        "viewport": parser.viewport,
        "schemaBlocks": parser.schema_blocks,
        "schemaTypes": sorted(parser.schema_types),
        "wordCount": parser.text_words,
        "internalLinks": len(internal_urls),
        "externalLinks": len(external_urls),
        "internalLinkSamples": internal_urls[:24],
        "externalLinkSamples": external_urls[:16],
        "socialLinks": list(dict.fromkeys([*sorted(parser.social_links), *contacts["socialLinks"]]))[:20],
        "phoneLinks": list(dict.fromkeys([*sorted(parser.phone_links), *contacts["phoneNumbers"]]))[:20],
        "emailLinks": list(dict.fromkeys([*sorted(parser.email_links), *contacts["emails"]]))[:20],
        "whatsappLinks": contacts["whatsappLinks"],
        "whatsappNumber": contacts["whatsappNumber"],
        "tags": contacts["tags"],
        "publicAddresses": contacts["addresses"],
        "contactPageCandidates": contacts["contactPageCandidates"],
        "textCharacters": parser.text_chars,
        "robots": robots,
        "sitemap": sitemap,
        "seoScore": score,
        "issues": issues,
        "wins": wins,
        "checkedAt": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "disclaimer": "Single-route public audit. Recheck timeouts and availability from the target market."
    }


def provider_status():
    """Return configuration state without exposing credentials."""
    providers = {
        "whatsapp": bool(os.getenv("WHATSAPP_TOKEN") and os.getenv("WHATSAPP_PHONE_NUMBER_ID")),
        "telegram": bool(os.getenv("TELEGRAM_BOT_TOKEN")),
        "bale": bool(os.getenv("BALE_BOT_TOKEN")),
        "rubika": bool(os.getenv("RUBIKA_BOT_TOKEN")),
        "soroush": bool(os.getenv("SOROUSH_PARTNER_WEBHOOK_URL")),
        "eitaa": bool(os.getenv("EITAA_APP_TOKEN")),
        "email": bool(os.getenv("SMTP_HOST") and os.getenv("SMTP_FROM")),
        "sms": bool(os.getenv("SMS_WEBHOOK_URL")),
        "divar": bool(os.getenv("DIVAR_PARTNER_WEBHOOK_URL")),
    }
    divar_slug = re.sub(r"[^a-zA-Z0-9_-]", "", os.getenv("DIVAR_APP_SLUG", ""))
    database = supabase_settings()
    webhook_database = bool(os.getenv("LEAD_DATABASE_WEBHOOK_URL") or os.getenv("LEAD_INGEST_WEBHOOK_URL"))
    return {
        "ok": True,
        "sendEnabled": SEND_ENABLED,
        "dryRun": DRY_RUN,
        "providers": providers,
        "vendorSearchConfigured": bool(os.getenv("VENDOR_SEARCH_WEBHOOK_URL")),
        "clinicSearchConfigured": bool(os.getenv("CLINIC_SEARCH_WEBHOOK_URL") or os.getenv("BRAVE_SEARCH_API_KEY") or os.getenv("GOOGLE_PLACES_API_KEY") or os.getenv("GOOGLE_MAPS_API_KEY")),
        "clinicSearchProviders": {
            "webhook": bool(os.getenv("CLINIC_SEARCH_WEBHOOK_URL")),
            "googlePlaces": bool(os.getenv("GOOGLE_PLACES_API_KEY") or os.getenv("GOOGLE_MAPS_API_KEY")),
            "brave": bool(os.getenv("BRAVE_SEARCH_API_KEY")),
        },
        "geminiConfigured": bool(get_gemini_keys()),
        "scraperConfigured": bool(scraper_allowed_domains()),
        "contactEnrichmentEnabled": True,
        "contactEnrichmentMaxPages": max(1, min(int(os.getenv("CONTACT_ENRICH_MAX_PAGES", "3")), 5)),
        "leadDatabaseConfigured": bool(database["configured"] or webhook_database),
        "leadDatabaseProvider": "supabase" if database["configured"] else "webhook" if webhook_database else "none",
        "leadDatabaseTable": database["table"],
        "leadDatabaseDetectedVariables": {"url": database["urlVariable"], "key": database["keyVariable"]},
        "geminiModel": os.getenv("GEMINI_MODEL", "gemini-flash-lite-latest"),
        "proposalPdfMode": "direct-download" if PILLOW_AVAILABLE else "browser-print",
        "pdfTextEngine": "raqm" if RAQM_AVAILABLE else "arabic-reshaper+bidi" if BIDI_FALLBACK_AVAILABLE else "basic",
        "pdfLinkTtlSeconds": PDF_LINK_TTL,
        "pdfLinksEphemeral": True,
        "webApps": {
            "bale": "https://web.bale.ai",
            "rubika": "https://web.rubika.ir",
            "soroush": "https://web.splus.ir",
            "eitaa": "https://web.eitaa.com",
            "divar": f"https://divar.ir/chat/addon_{divar_slug}" if divar_slug else "https://divar.ir/",
        },
        "notes": {
            "whatsapp": "Official Meta Cloud API; opt-in and template/session rules apply.",
            "telegram": "The user must start the bot first, or the bot must have channel/group permission.",
            "bale": "Official Bale Bot API; the user must start the bot or authorize the conversation.",
            "rubika": "Official Rubika Bot API v3; chat_id and bot authorization are required.",
            "soroush": "Uses an operator-authorized Soroush Plus partner webhook.",
            "eitaa": "Uses the Eitaa application sendMessage API; token and permitted chat_id are required.",
            "email": "SMTP credentials remain server-side.",
            "sms": "Uses an approved provider webhook configured by the operator.",
            "divar": "Automatic sending is available only through an authorized Divar partner webhook; no scraping or browser automation.",
        },
    }


def rate_limit(channel: str, recipient: str):
    """Small-process safety limit: 5 sends/minute per channel+recipient, 30/minute total."""
    now = time.monotonic()
    keys = [f"recipient:{channel}:{hashlib.sha256(recipient.encode()).hexdigest()[:16]}", "global"]
    limits = [5, 30]
    for key, limit in zip(keys, limits):
        bucket = RATE_BUCKETS[key]
        while bucket and now - bucket[0] > 60:
            bucket.popleft()
        if len(bucket) >= limit:
            raise ValueError("Rate limit reached. Wait before sending again.")
    for key in keys:
        RATE_BUCKETS[key].append(now)


def get_gemini_keys():
    keys = []
    for index in range(1, 4):
        value = os.getenv(f"GEMINI_API_KEY{index}", "").strip()
        if value and value not in keys:
            keys.append(value)
    fallback = os.getenv("GEMINI_API_KEY", "").strip()
    if fallback and fallback not in keys:
        keys.append(fallback)
    return keys


def call_gemini(prompt: str, temperature: float = 0.65, max_tokens: int = 12000):
    keys = get_gemini_keys()
    if not keys:
        raise ValueError("No Gemini API key is configured. Add GEMINI_API_KEY1 or GEMINI_API_KEY.")
    model = os.getenv("GEMINI_MODEL", "gemini-flash-lite-latest").strip()
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent"
    config = {"maxOutputTokens": max(1000, min(max_tokens, 16000)), "temperature": temperature}
    if os.getenv("GEMINI_USE_THINKING", "false").lower() == "true":
        config["thinkingConfig"] = {"thinkingLevel": "low"}
    payload = {"contents": [{"parts": [{"text": prompt}]}], "generationConfig": config}
    errors = []
    for number, key in enumerate(keys, 1):
        try:
            response = requests.post(url, headers={"x-goog-api-key": key, "Content-Type": "application/json"},
                                     json=payload, timeout=120)
            data = response.json() if response.text else {}
            if not response.ok:
                error = data.get("error", {}) if isinstance(data, dict) else {}
                status = str(error.get("status", ""))
                message = str(error.get("message", response.text or "Unknown Gemini error"))
                if response.status_code == 429 or status == "RESOURCE_EXHAUSTED" or "quota" in message.lower():
                    errors.append(f"key {number}: quota exhausted")
                    continue
                raise ValueError(f"Gemini API: {message[:500]}")
            candidates = data.get("candidates") or []
            if not candidates:
                raise ValueError("Gemini returned no candidate.")
            parts = ((candidates[0].get("content") or {}).get("parts") or [])
            text = "\n".join(str(part.get("text", "")) for part in parts
                             if isinstance(part, dict) and not part.get("thought") and part.get("text"))
            if not text.strip():
                raise ValueError("Gemini returned an empty article.")
            return text.strip(), number, model
        except requests.RequestException as exc:
            errors.append(f"key {number}: {type(exc).__name__}")
            continue
    raise ValueError("All Gemini keys failed or reached quota: " + "; ".join(errors))


def split_seo_keywords(raw: str):
    return [item.strip() for item in re.split(r"،|,|\s+-\s+", raw or "") if item.strip()][:12]


def keyword_occurrences(text: str, keyword: str):
    if not keyword:
        return 0
    return len(re.findall(re.escape(keyword), text, flags=re.IGNORECASE))


def build_article_prompt(data: dict, target: int, secondary: list[str], min_primary: int, max_primary: int,
                         min_secondary: int):
    language = data["language"]
    title = data["title"]
    outline = data["outline"]
    primary = data["primaryKeyword"]
    secondary_text = "، ".join(secondary) if secondary else "ندارد"
    rewrite = data.get("isRewrite") is True
    notes = str(data.get("rewriteNotes", "")).strip()[:1000]
    if language == "en":
        prompt = f"""You are a senior human SEO editor. Write a useful, natural, publication-ready article in English.

Title: {title}
Primary keyword: {primary}
Secondary keywords: {', '.join(secondary) if secondary else 'none'}
Required H2 outline (use every line exactly, in this exact order, without adding or rewriting headings):
{outline}

Requirements:
- Target {target} words, tolerance ±30 words.
- Use the exact primary keyword {min_primary} to {max_primary} times, naturally distributed, including once in the first 100 words.
- Use each secondary keyword at least {min_secondary} times when contextually relevant; never stuff keywords.
- After the introduction, add an H2 named exactly “Quick overview 👀”, two short sentences and a bullet list.
- For each supplied H2: a short introduction, useful main content, and a natural transition. Use H3 only where the H2 contains multiple distinct ideas.
- Prefer active voice, direct reader address, specific examples and practical guidance.
- Do not invent statistics, medical claims, certifications, prices, testimonials or sources.
- For medical topics, provide general educational information, avoid diagnosis or treatment promises, and add a short “Medical review required” note.
- Do not output an SEO report, analysis, preface, or code fence. Output only the article in Markdown.
"""
    else:
        prompt = f"""تو یک ویراستار حرفه‌ای و انسانی سئو هستی. یک مقاله کاربردی، طبیعی و آماده انتشار به زبان فارسی معیار بنویس.

عنوان: {title}
کلمه کلیدی اصلی: {primary}
کلمات کلیدی فرعی: {secondary_text}
Outline اجباری H2 (هر خط را دقیقاً با همین متن و همین ترتیب استفاده کن و عنوانی را تغییر نده):
{outline}

قواعد:
- طول هدف {target} کلمه با تلورانس حداکثر ±۳۰ کلمه.
- عبارت دقیق «{primary}» را بین {min_primary} تا {max_primary} بار، طبیعی و توزیع‌شده استفاده کن و یک بار در ۱۰۰ کلمه اول بیاور.
- هر کلمه فرعی مرتبط را حداقل {min_secondary} بار طبیعی استفاده کن؛ حشو کلمه ممنوع است.
- بعد از مقدمه یک H2 با عنوان دقیق «نگاه سریع 👀» شامل دو جمله کوتاه و یک فهرست نقطه‌ای اضافه کن.
- برای هر H2 داده‌شده: مقدمه کوتاه، محتوای ارزشمند و گذار طبیعی. فقط برای چند ایده مستقل H3 بساز.
- از لحن مستقیم، فعل معلوم، مثال مشخص و راهکار عملی استفاده کن.
- آمار، ادعای پزشکی، مجوز، قیمت، رضایت مشتری یا منبع ساختگی تولید نکن.
- در موضوعات پزشکی فقط اطلاعات آموزشی عمومی بده، تشخیص یا وعده درمان ارائه نکن و در پایان یادداشت کوتاه «نیازمند بازبینی پزشک» اضافه کن.
- گزارش سئو، توضیح فرایند، مقدمه خارج از مقاله یا code fence تولید نکن. فقط مقاله Markdown را برگردان.
"""
    if rewrite:
        prompt += (f"\nRewrite the article from scratch with a substantially different expression. Address this note: {notes or 'Improve clarity and naturalness'}.\n"
                   if language == "en" else
                   f"\nمقاله را از ابتدا با بیان کاملاً متفاوت بازنویسی کن و این ملاحظه را اعمال کن: {notes or 'شفافیت و طبیعی‌بودن متن بهتر شود'}.\n")
    return prompt


def article_report(article: str, data: dict, target: int, secondary: list[str]):
    words = [word for word in article.split() if word]
    word_count = len(words)
    primary = data["primaryKeyword"]
    primary_count = keyword_occurrences(article, primary)
    min_primary = math.ceil(max(word_count, 1) * 0.01)
    max_primary = max(min_primary, math.floor(max(word_count, 1) * 0.015))
    first_100 = " ".join(words[:100])
    report = {
        "wordCount": word_count,
        "targetWordCount": target,
        "wordCountPass": abs(word_count - target) <= 30,
        "primaryKeyword": primary,
        "primaryCount": primary_count,
        "primaryMin": min_primary,
        "primaryMax": max_primary,
        "density": round(primary_count / max(word_count, 1) * 100, 2),
        "primaryInFirst100": keyword_occurrences(first_100, primary) > 0,
        "secondary": [{"keyword": keyword, "count": keyword_occurrences(article, keyword)} for keyword in secondary],
    }
    return report


def generate_seo_article(payload: dict):
    language = str(payload.get("language", "fa")).lower()
    language = "en" if language == "en" else "fa"
    data = {
        "language": language,
        "title": str(payload.get("title", "")).strip()[:300],
        "outline": str(payload.get("outline", "")).strip()[:5000],
        "primaryKeyword": str(payload.get("primaryKeyword", "")).strip()[:250],
        "isRewrite": payload.get("isRewrite") is True,
        "rewriteNotes": str(payload.get("rewriteNotes", "")).strip()[:1000],
    }
    if not data["title"] or not data["outline"] or not data["primaryKeyword"]:
        raise ValueError("Title, outline and primary keyword are required.")
    try:
        target = int(payload.get("targetWordCount", 900))
    except (TypeError, ValueError):
        target = 900
    target = max(800, min(target, 1500))
    secondary = split_seo_keywords(str(payload.get("secondaryKeywords", "")))
    min_primary = math.ceil(target * 0.01)
    max_primary = max(min_primary, math.floor(target * 0.015))
    min_secondary = max(2, round(target / 900 * 3))
    prompt = build_article_prompt(data, target, secondary, min_primary, max_primary, min_secondary)
    article, key_number, model = call_gemini(prompt)
    article = re.sub(r"\n\s*\*\*\*\s*\n\s*#{2,3}\s*(SEO|سئو).*$", "", article, flags=re.IGNORECASE | re.DOTALL).strip()
    report = article_report(article, data, target, secondary)
    needs_correction = (abs(report["wordCount"] - target) > 70 or
                        report["primaryCount"] < report["primaryMin"] or
                        report["primaryCount"] > report["primaryMax"])
    corrected = False
    if needs_correction and os.getenv("GEMINI_AUTO_CORRECT", "true").lower() != "false":
        if language == "en":
            correction = f"""Revise the Markdown article below. Keep every existing H2 heading exactly unchanged. Reach {target} words ±30. Use the exact primary keyword “{data['primaryKeyword']}” between {min_primary} and {max_primary} times, naturally, and once in the first 100 words. Preserve factual caution and do not add fabricated claims. Return only the complete revised article.\n\n{article}"""
        else:
            correction = f"""مقاله Markdown زیر را اصلاح کن. تمام H2های موجود را دقیقاً بدون تغییر نگه دار. متن را به {target} کلمه با تلورانس ±۳۰ برسان. عبارت دقیق «{data['primaryKeyword']}» را بین {min_primary} تا {max_primary} بار طبیعی و یک بار در ۱۰۰ کلمه اول استفاده کن. احتیاط علمی را حفظ کن و ادعای ساختگی نساز. فقط متن کامل اصلاح‌شده را برگردان.\n\n{article}"""
        try:
            article, key_number, model = call_gemini(correction, temperature=0.45)
            report = article_report(article, data, target, secondary)
            corrected = True
        except Exception:
            corrected = False
    return {"ok": True, "article": article, "report": report, "language": language,
            "model": model, "keyNumber": key_number, "autoCorrected": corrected,
            "disclaimer": "AI-generated content requires human editorial review; medical content also requires qualified medical review."}


def parse_ai_json(text: str):
    cleaned = text.strip()
    cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s*```$", "", cleaned)
    start, end = cleaned.find("{"), cleaned.rfind("}")
    if start < 0 or end <= start:
        raise ValueError("Gemini did not return a JSON object.")
    try:
        return json.loads(cleaned[start:end + 1])
    except json.JSONDecodeError as exc:
        raise ValueError(f"Gemini returned invalid JSON: {exc.msg}") from exc


def generate_ai_seo_review(payload: dict):
    language = "en" if str(payload.get("language", "fa")).lower() == "en" else "fa"
    lead = payload.get("lead") if isinstance(payload.get("lead"), dict) else {}
    supplied_audit = payload.get("audit") if isinstance(payload.get("audit"), dict) else None
    url = str(payload.get("url", "")).strip()
    if supplied_audit:
        measured = supplied_audit
    elif url:
        measured = audit(url)
    else:
        raise ValueError("A public URL or measured audit object is required.")

    evidence = {
        "requestedUrl": str(measured.get("requestedUrl", url))[:500],
        "finalUrl": str(measured.get("finalUrl", ""))[:500],
        "httpStatus": measured.get("status"),
        "measuredSeoScore": measured.get("seoScore"),
        "elapsedSeconds": measured.get("elapsedSeconds"),
        "title": str(measured.get("title", ""))[:300],
        "titleLength": measured.get("titleLength"),
        "description": str(measured.get("description", ""))[:600],
        "descriptionLength": measured.get("descriptionLength"),
        "h1Count": measured.get("h1Count"),
        "h1": measured.get("h1", [])[:5] if isinstance(measured.get("h1"), list) else [],
        "canonical": str(measured.get("canonical", ""))[:500],
        "schemaTypes": measured.get("schemaTypes", [])[:30] if isinstance(measured.get("schemaTypes"), list) else [],
        "internalLinks": measured.get("internalLinks"),
        "externalLinks": measured.get("externalLinks"),
        "internalLinkSamples": measured.get("internalLinkSamples", [])[:12] if isinstance(measured.get("internalLinkSamples"), list) else [],
        "socialLinks": measured.get("socialLinks", [])[:10] if isinstance(measured.get("socialLinks"), list) else [],
        "phoneLinks": measured.get("phoneLinks", [])[:8] if isinstance(measured.get("phoneLinks"), list) else [],
        "emailLinks": measured.get("emailLinks", [])[:8] if isinstance(measured.get("emailLinks"), list) else [],
        "textCharacters": measured.get("textCharacters"),
        "robots": measured.get("robots"),
        "sitemap": measured.get("sitemap"),
        "issues": measured.get("issues", [])[:12] if isinstance(measured.get("issues"), list) else [],
        "wins": measured.get("wins", [])[:10] if isinstance(measured.get("wins"), list) else [],
    }
    lead_context = {
        "name": str(lead.get("name", ""))[:180],
        "publicScale": str(lead.get("scale", ""))[:30],
        "area": str(lead.get("area", ""))[:300],
        "services": str(lead.get("services", ""))[:500],
        "existingOpportunityScore": lead.get("opportunity"),
    }
    output_language = "English" if language == "en" else "Persian"
    prompt = f"""You are a senior technical SEO strategist for medical-clinic websites. Analyze only the measured evidence below and return valid JSON, with every human-readable value in {output_language}.

MEASURED AUDIT EVIDENCE:
{json.dumps(evidence, ensure_ascii=False, indent=2)}

PUBLIC LEAD CONTEXT (may be incomplete and must not be treated as verified revenue data):
{json.dumps(lead_context, ensure_ascii=False, indent=2)}

Rules:
- Separate measured facts from AI interpretation.
- Never invent Google positions, traffic, backlinks, revenue, patient numbers, licenses, reviews or medical outcomes.
- Never guarantee rank 1 or a treatment result.
- Treat publicScale as a rough sales segmentation label, not income.
- Medical content recommendations require qualified medical review.
- Prioritize fixes by evidence, impact and effort.
- opportunityScore is an advisory sales-fit score, not a factual business valuation.
- Budget ranges are editable planning estimates in Iranian toman.
- Outreach must be respectful, mention one evidence-based observation, ask permission to send a report and include a no-more-messages option.
- Return JSON only, with no Markdown fence.

Required JSON schema:
{{
  "executiveSummary": "string",
  "aiSeoScore": 0,
  "opportunityScore": 0,
  "confidence": "low|medium|high",
  "measuredFacts": ["string"],
  "issues": [{{"title":"string","severity":"critical|high|medium|low","evidence":"string","impact":"string","fix":"string","effort":"small|medium|large"}}],
  "quickWins": ["string"],
  "contentGaps": [{{"cluster":"string","intent":"commercial|informational|local","recommendedAssets":["string"]}}],
  "roadmap": {{"days1to30":["string"],"days31to60":["string"],"days61to90":["string"]}},
  "package": {{"name":"string","setupBudget":"string","monthlyFee":"string","mediaBudget":"string","duration":"string","reason":"string"}},
  "kpis": ["string"],
  "risksAndAssumptions": ["string"],
  "outreach": {{"whatsapp":"string","emailSubject":"string","emailBody":"string"}}
}}
"""
    raw, key_number, model = call_gemini(prompt, temperature=0.25, max_tokens=9000)
    analysis = parse_ai_json(raw)
    for score_key in ("aiSeoScore", "opportunityScore"):
        try:
            analysis[score_key] = max(0, min(100, int(analysis.get(score_key, 0))))
        except (TypeError, ValueError):
            analysis[score_key] = 0
    if analysis.get("confidence") not in {"low", "medium", "high"}:
        analysis["confidence"] = "low"
    return {
        "ok": True,
        "language": language,
        "measuredAudit": measured,
        "aiAnalysis": analysis,
        "model": model,
        "keyNumber": key_number,
        "disclaimer": "Measured audit fields are deterministic observations. AI scores and recommendations are advisory and require human validation."
    }


def analyze_clinic_candidates_ai(payload: dict):
    items = payload.get("items") if isinstance(payload.get("items"), list) else []
    if not items:
        raise ValueError("Select at least one clinic candidate for AI analysis.")
    language = "en" if str(payload.get("language", "fa")).lower() == "en" else "fa"
    compact = []
    for index, item in enumerate(items[:20]):
        if not isinstance(item, dict):
            continue
        compact.append({"index": index, "name": str(item.get("name", ""))[:220],
                        "website": str(item.get("website", ""))[:500],
                        "summary": str(item.get("summary", ""))[:700],
                        "currentType": str(item.get("resultType", "web-result"))[:80],
                        "specialty": str(item.get("specialty", ""))[:150],
                        "phone": str(item.get("phone", ""))[:100],
                        "address": str(item.get("address", ""))[:300]})
    output_language = "English" if language == "en" else "Persian"
    prompt = f"""You classify public web-search results for a medical-clinic lead database. Return JSON only. Write all explanatory strings in {output_language}.

Candidates:
{json.dumps(compact, ensure_ascii=False, indent=2)}

Rules:
- Do not claim a result is licensed, active or official without evidence.
- Distinguish an actual clinic/physician profile from a list article, price article, directory page or unrelated page.
- Normalize the display name by removing domains, URLs, breadcrumbs, year labels, marketing symbols and generic list prefixes.
- Do not infer patient traits, health conditions, income or medical outcomes.
- confidence is advisory based only on supplied title/URL/snippet.
- priority is for verification workflow, not medical quality.
- Return one result for every input index.

Schema:
{{"items":[{{"index":0,"normalizedName":"string","isLikelyMedicalClinic":true,"resultType":"official-clinic|physician-profile|directory-profile|list-article|price-article|unrelated|uncertain","specialty":"string","confidence":0,"priority":"high|medium|low","reason":"string","recommendedNextStep":"string"}}]}}
"""
    raw, key_number, model = call_gemini(prompt, temperature=0.15, max_tokens=7000)
    parsed = parse_ai_json(raw)
    results = parsed.get("items") if isinstance(parsed.get("items"), list) else []
    output = []
    valid_types = {"official-clinic", "physician-profile", "directory-profile", "list-article", "price-article", "unrelated", "uncertain"}
    for result in results[:20]:
        if not isinstance(result, dict):
            continue
        try:
            index = int(result.get("index"))
        except (TypeError, ValueError):
            continue
        if index < 0 or index >= len(compact):
            continue
        try:
            confidence = max(0, min(100, int(result.get("confidence", 0))))
        except (TypeError, ValueError):
            confidence = 0
        result_type = str(result.get("resultType", "uncertain"))
        if result_type not in valid_types:
            result_type = "uncertain"
        output.append({"index": index, "normalizedName": str(result.get("normalizedName", compact[index]["name"]))[:180],
                       "isLikelyMedicalClinic": bool(result.get("isLikelyMedicalClinic", False)),
                       "resultType": result_type, "specialty": str(result.get("specialty", compact[index]["specialty"]))[:150],
                       "confidence": confidence, "priority": str(result.get("priority", "medium"))[:20],
                       "reason": str(result.get("reason", ""))[:500],
                       "recommendedNextStep": str(result.get("recommendedNextStep", ""))[:500]})
    return {"ok": True, "items": output, "model": model, "keyNumber": key_number,
            "disclaimer": "AI classification is advisory. Verify identity, medical license, official ownership and contact data independently."}


def post_json(url: str, payload: dict, headers: dict | None = None, timeout: int = 20):
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    hdr = {"User-Agent": USER_AGENT, "Content-Type": "application/json", "Accept": "application/json"}
    if headers:
        hdr.update(headers)
    req = Request(url, data=body, headers=hdr, method="POST")
    opener = build_opener(SafeRedirect())
    try:
        with opener.open(req, timeout=timeout) as response:
            raw = response.read(500_000).decode("utf-8", errors="replace")
            try:
                data = json.loads(raw)
            except Exception:
                data = {"raw": raw[:1000]}
            return int(response.status), data
    except HTTPError as exc:
        raw = exc.read(200_000).decode("utf-8", errors="replace")
        raise ValueError(f"Provider HTTP {exc.code}: {raw[:500]}")


def get_json(url: str, headers: dict | None = None, timeout: int = 20):
    hdr = {"User-Agent": USER_AGENT, "Accept": "application/json"}
    if headers:
        hdr.update(headers)
    req = Request(url, headers=hdr, method="GET")
    opener = build_opener(SafeRedirect())
    try:
        with opener.open(req, timeout=timeout) as response:
            raw = response.read(1_000_000).decode("utf-8", errors="replace")
            return int(response.status), json.loads(raw)
    except HTTPError as exc:
        raw = exc.read(200_000).decode("utf-8", errors="replace")
        raise ValueError(f"Search provider HTTP {exc.code}: {raw[:500]}")


def post_multipart(url: str, fields: dict, file_field: str, filename: str, content_type: str,
                   file_bytes: bytes, headers: dict | None = None, timeout: int = 30):
    boundary = "----ClinicSignal" + hashlib.sha256(os.urandom(16)).hexdigest()[:24]
    chunks: list[bytes] = []
    for key, value in fields.items():
        chunks.extend([f"--{boundary}\r\n".encode(),
                       f'Content-Disposition: form-data; name="{key}"\r\n\r\n'.encode(),
                       str(value).encode("utf-8"), b"\r\n"])
    chunks.extend([f"--{boundary}\r\n".encode(),
                   f'Content-Disposition: form-data; name="{file_field}"; filename="{filename}"\r\n'.encode(),
                   f"Content-Type: {content_type}\r\n\r\n".encode(), file_bytes, b"\r\n",
                   f"--{boundary}--\r\n".encode()])
    body = b"".join(chunks)
    hdr = {"User-Agent": USER_AGENT, "Accept": "application/json",
           "Content-Type": f"multipart/form-data; boundary={boundary}"}
    if headers:
        hdr.update(headers)
    req = Request(url, data=body, headers=hdr, method="POST")
    opener = build_opener(SafeRedirect())
    try:
        with opener.open(req, timeout=timeout) as response:
            raw = response.read(500_000).decode("utf-8", errors="replace")
            try:
                data = json.loads(raw)
            except Exception:
                data = {"raw": raw[:1000]}
            return int(response.status), data
    except HTTPError as exc:
        raw = exc.read(200_000).decode("utf-8", errors="replace")
        raise ValueError(f"Provider HTTP {exc.code}: {raw[:500]}")


def send_email(recipient: str, message: str, subject: str, attachment: bytes | None = None,
               attachment_name: str = "proposal.pdf"): 
    host = os.getenv("SMTP_HOST")
    sender = os.getenv("SMTP_FROM")
    if not host or not sender:
        raise ValueError("Email provider is not configured.")
    port = int(os.getenv("SMTP_PORT", "587"))
    user = os.getenv("SMTP_USER", "")
    password = os.getenv("SMTP_PASSWORD", "")
    use_ssl = os.getenv("SMTP_SSL", "false").lower() == "true"
    email = EmailMessage()
    email["From"] = sender
    email["To"] = recipient
    email["Subject"] = subject or "Clinic Signal message"
    email.set_content(message)
    if attachment:
        email.add_attachment(attachment, maintype="application", subtype="pdf", filename=attachment_name)
    if use_ssl:
        smtp = smtplib.SMTP_SSL(host, port, timeout=20, context=ssl.create_default_context())
    else:
        smtp = smtplib.SMTP(host, port, timeout=20)
        if os.getenv("SMTP_STARTTLS", "true").lower() != "false":
            smtp.starttls(context=ssl.create_default_context())
    try:
        if user:
            smtp.login(user, password)
        smtp.send_message(email)
    finally:
        smtp.quit()
    return {"accepted": True}


def send_message(payload: dict):
    channel = str(payload.get("channel", "")).lower().strip()
    recipient = str(payload.get("recipient", "")).strip()
    message = str(payload.get("message", "")).strip()
    subject = str(payload.get("subject", "")).strip()
    if channel not in ALLOWED_CHANNELS:
        raise ValueError("Unsupported channel.")
    if not recipient or not message:
        raise ValueError("Recipient and message are required.")
    if len(message) > 4000:
        raise ValueError("Message is longer than 4000 characters.")
    if payload.get("approved") is not True:
        raise ValueError("Human approval is required before sending.")
    if payload.get("consent") is not True:
        raise ValueError("Documented recipient consent or an existing service conversation is required.")
    if payload.get("senderAuthorized") is not True:
        raise ValueError("Authorization to represent the selected sender company is required.")
    if payload.get("doNotContact") is True:
        raise ValueError("Recipient is on the do-not-contact list.")
    rate_limit(channel, recipient)

    recipient_hash = hashlib.sha256(recipient.encode("utf-8")).hexdigest()[:16]
    log = {"time": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()), "channel": channel,
           "recipientHash": recipient_hash, "leadId": str(payload.get("leadId", ""))[:80]}
    pdf_bytes = None
    pdf_filename = "proposal.pdf"
    requested_attachment = payload.get("attachProposalPdf") is True
    attach_pdf = requested_attachment and channel in {"whatsapp", "telegram", "bale", "email"}
    manual_attachment_required = requested_attachment and not attach_pdf
    if attach_pdf:
        proposal = payload.get("proposal") if isinstance(payload.get("proposal"), dict) else None
        if not proposal:
            raise ValueError("Proposal data is required for a PDF attachment.")
        pdf_bytes, pdf_filename = make_proposal_pdf(proposal)

    if DRY_RUN or not SEND_ENABLED:
        log.update(status="simulated", attachment=bool(pdf_bytes))
        SEND_LOG.appendleft(log)
        return {"ok": True, "sent": False, "dryRun": True, "status": "simulated",
                "attachmentReady": bool(pdf_bytes), "manualAttachmentRequired": manual_attachment_required,
                "message": "Validated successfully. Sending is disabled or DRY_RUN is active."}

    if manual_attachment_required:
        raise ValueError("Automatic PDF attachment is not configured for this channel. Download the PDF and use the official web app handoff.")
    configured = provider_status()["providers"]
    if not configured[channel]:
        raise ValueError(f"{channel.title()} provider is not configured.")

    if channel == "whatsapp":
        token = os.environ["WHATSAPP_TOKEN"]
        phone_id = os.environ["WHATSAPP_PHONE_NUMBER_ID"]
        version = os.getenv("WHATSAPP_API_VERSION", "v23.0")
        url = f"https://graph.facebook.com/{version}/{phone_id}/messages"
        auth = {"Authorization": f"Bearer {token}"}
        status, text_response = post_json(url, {"messaging_product": "whatsapp", "to": re.sub(r"\D", "", recipient),
                    "type": "text", "text": {"preview_url": False, "body": message}}, auth)
        response = {"text": text_response}
        if pdf_bytes:
            media_url = f"https://graph.facebook.com/{version}/{phone_id}/media"
            _, media_response = post_multipart(media_url, {"messaging_product": "whatsapp"}, "file",
                                                pdf_filename, "application/pdf", pdf_bytes, auth)
            media_id = media_response.get("id") if isinstance(media_response, dict) else None
            if not media_id:
                raise ValueError("WhatsApp media upload did not return a media id.")
            doc_status, doc_response = post_json(url, {"messaging_product": "whatsapp",
                "to": re.sub(r"\D", "", recipient), "type": "document",
                "document": {"id": media_id, "filename": pdf_filename}}, auth)
            status = doc_status
            response["document"] = doc_response
    elif channel == "telegram":
        token = os.environ["TELEGRAM_BOT_TOKEN"]
        url = f"https://api.telegram.org/bot{token}/sendMessage"
        status, text_response = post_json(url, {"chat_id": recipient, "text": message, "disable_web_page_preview": True})
        response = {"text": text_response}
        if pdf_bytes:
            doc_url = f"https://api.telegram.org/bot{token}/sendDocument"
            doc_status, doc_response = post_multipart(doc_url, {"chat_id": recipient}, "document",
                                                       pdf_filename, "application/pdf", pdf_bytes)
            status = doc_status
            response["document"] = doc_response
    elif channel == "bale":
        token = os.environ["BALE_BOT_TOKEN"]
        base = f"https://tapi.bale.ai/bot{token}"
        status, text_response = post_json(base + "/sendMessage", {"chat_id": recipient, "text": message})
        response = {"text": text_response}
        if pdf_bytes:
            doc_status, doc_response = post_multipart(base + "/sendDocument", {"chat_id": recipient},
                                                       "document", pdf_filename, "application/pdf", pdf_bytes)
            status = doc_status
            response["document"] = doc_response
    elif channel == "rubika":
        token = os.environ["RUBIKA_BOT_TOKEN"]
        url = f"https://botapi.rubika.ir/v3/{token}/sendMessage"
        status, response = post_json(url, {"chat_id": recipient, "text": message})
    elif channel == "soroush":
        url = os.environ["SOROUSH_PARTNER_WEBHOOK_URL"]
        token = os.getenv("SOROUSH_PARTNER_TOKEN", "")
        headers = {"Authorization": f"Bearer {token}"} if token else {}
        status, response = post_json(url, {"recipient": recipient, "message": message}, headers)
    elif channel == "eitaa":
        status, response = post_json("https://eitaayar.ir/api/app/sendMessage", {
            "token": os.environ["EITAA_APP_TOKEN"], "chat_id": recipient, "text": message})
    elif channel == "email":
        response = send_email(recipient, message, subject, pdf_bytes, pdf_filename)
        status = 200
    elif channel == "sms":
        url = os.environ["SMS_WEBHOOK_URL"]
        headers = {"Authorization": f"Bearer {os.getenv('SMS_WEBHOOK_TOKEN', '')}"} if os.getenv("SMS_WEBHOOK_TOKEN") else {}
        status, response = post_json(url, {"to": recipient, "message": message,
                                            "sender": os.getenv("SMS_SENDER", "")}, headers)
    elif channel == "divar":  # Authorized Kenar-e-Divar middleware only
        url = os.environ["DIVAR_PARTNER_WEBHOOK_URL"]
        headers = {"Authorization": f"Bearer {os.getenv('DIVAR_PARTNER_TOKEN', '')}"} if os.getenv("DIVAR_PARTNER_TOKEN") else {}
        status, response = post_json(url, {"conversation_id": recipient, "message": message,
                                           "app_slug": os.getenv("DIVAR_APP_SLUG", "")}, headers)
    else:
        raise ValueError("Channel adapter is not implemented.")

    log.update(status="sent", providerStatus=status)
    SEND_LOG.appendleft(log)
    return {"ok": True, "sent": True, "dryRun": False, "status": "sent", "providerStatus": status,
            "providerResponse": response}


def search_vendors(payload: dict):
    """Use an operator-configured public-search adapter; never scrape search engines directly."""
    query = str(payload.get("query", "")).strip()[:300]
    location = str(payload.get("location", "Tehran")).strip()[:100]
    categories = payload.get("categories") or []
    if not query:
        raise ValueError("Search query is required.")
    if not isinstance(categories, list):
        categories = []
    categories = [str(x)[:80] for x in categories[:8]]
    webhook = os.getenv("VENDOR_SEARCH_WEBHOOK_URL", "")
    token = os.getenv("VENDOR_SEARCH_WEBHOOK_TOKEN", "")
    if webhook:
        headers = {"Authorization": f"Bearer {token}"} if token else {}
        status, response = post_json(webhook, {"query": query, "location": location,
                                               "categories": categories, "limit": 12}, headers)
        items = response.get("items", []) if isinstance(response, dict) else []
        clean = []
        for item in items[:12]:
            if not isinstance(item, dict):
                continue
            clean.append({
                "name": str(item.get("name", ""))[:160],
                "website": str(item.get("website", ""))[:500],
                "category": str(item.get("category", ""))[:100],
                "location": str(item.get("location", ""))[:120],
                "evidence": str(item.get("evidence", item.get("source", "")))[:500],
                "phone": str(item.get("phone", ""))[:80],
                "summary": str(item.get("summary", ""))[:500],
                "verified": bool(item.get("verified", False)),
            })
        return {"ok": True, "configured": True, "providerStatus": status, "items": clean,
                "disclaimer": "Search results are candidates, not endorsements. Verify scope, references and credentials."}
    text = f"{query} {location}".strip()
    return {
        "ok": True,
        "configured": False,
        "items": [],
        "searchLinks": {
            "google": "https://www.google.com/search?q=" + quote_plus(text),
            "linkedin": "https://www.linkedin.com/search/results/companies/?keywords=" + quote_plus(text),
        },
        "disclaimer": "No search adapter is configured. Use the generated public-search links or add candidates manually.",
    }


def search_clinics(payload: dict):
    """Search adapter for public medical-clinic discovery; never collects patient data."""
    query = str(payload.get("query", "")).strip()[:350]
    location = str(payload.get("location", "Tehran")).strip()[:120]
    specialty = str(payload.get("specialty", "medical clinic")).strip()[:120]
    engines = payload.get("engines") if isinstance(payload.get("engines"), list) else []
    engines = [str(x).lower()[:30] for x in engines[:8]]
    if not query:
        query = f"{specialty} {location} official website contact"
    combined = f"{query} {location}".strip()
    webhook = os.getenv("CLINIC_SEARCH_WEBHOOK_URL", "")
    token = os.getenv("CLINIC_SEARCH_WEBHOOK_TOKEN", "")
    if webhook:
        headers = {"Authorization": f"Bearer {token}"} if token else {}
        status, response = post_json(webhook, {
            "query": query, "location": location, "specialty": specialty,
            "engines": engines, "limit": 20, "publicBusinessOnly": True,
        }, headers)
        source_items = response.get("items", []) if isinstance(response, dict) else []
        items = []
        for item in source_items[:20]:
            if not isinstance(item, dict):
                continue
            items.append({
                "name": str(item.get("name", ""))[:180],
                "website": str(item.get("website", ""))[:500],
                "phone": normalize_public_phone(str(item.get("phone", "")))[:100],
                "email": str(item.get("email", ""))[:180],
                "whatsapp": normalize_public_phone(str(item.get("whatsapp", "")))[:100],
                "whatsappLinks": [str(x)[:500] for x in item.get("whatsappLinks", [])[:10]] if isinstance(item.get("whatsappLinks"), list) else [],
                "address": str(item.get("address", ""))[:500],
                "specialty": str(item.get("specialty", specialty))[:150],
                "tags": [str(x)[:80] for x in item.get("tags", [])[:30]] if isinstance(item.get("tags"), list) else [],
                "source": str(item.get("source", item.get("evidence", "")))[:500],
                "summary": str(item.get("summary", ""))[:600],
                "verified": bool(item.get("verified", False)),
            })
        return {"ok": True, "configured": True, "mode": "api", "provider": "webhook", "providerStatus": status, "items": items,
                "disclaimer": "Candidates only. Verify medical license, identity, public contact details and active status independently."}
    google_places_key = os.getenv("GOOGLE_PLACES_API_KEY", "").strip() or os.getenv("GOOGLE_MAPS_API_KEY", "").strip()
    if google_places_key and (not engines or "google" in engines):
        field_mask = ",".join([
            "places.id", "places.displayName", "places.formattedAddress",
            "places.nationalPhoneNumber", "places.internationalPhoneNumber",
            "places.websiteUri", "places.googleMapsUri", "places.primaryTypeDisplayName",
            "places.businessStatus",
        ])
        status, response = post_json(
            "https://places.googleapis.com/v1/places:searchText",
            {"textQuery": combined, "languageCode": "fa", "pageSize": 20},
            {"X-Goog-Api-Key": google_places_key, "X-Goog-FieldMask": field_mask},
            timeout=25,
        )
        places = response.get("places", []) if isinstance(response, dict) else []
        items = []
        for place in places[:20]:
            if not isinstance(place, dict):
                continue
            display = place.get("displayName") if isinstance(place.get("displayName"), dict) else {}
            primary_type = place.get("primaryTypeDisplayName") if isinstance(place.get("primaryTypeDisplayName"), dict) else {}
            website = str(place.get("websiteUri", ""))[:500]
            maps_url = str(place.get("googleMapsUri", ""))[:500]
            business_status = str(place.get("businessStatus", ""))[:60]
            no_site = not bool(website)
            items.append({
                "name": str(display.get("text", ""))[:180],
                "website": website,
                "phone": normalize_public_phone(str(place.get("internationalPhoneNumber") or place.get("nationalPhoneNumber") or ""))[:100],
                "email": "", "whatsapp": "", "whatsappLinks": [],
                "address": str(place.get("formattedAddress", ""))[:500],
                "specialty": str(primary_type.get("text") or specialty)[:150],
                "tags": [str(primary_type.get("text") or specialty)[:80]],
                "source": maps_url,
                "summary": f"Google Places public business result. Status: {business_status or 'not provided'}",
                "verified": False,
                "resultType": "structured-medical-entity",
                "placeId": str(place.get("id", ""))[:200],
                "websiteStatus": "no-website-found" if no_site else "official-website-provided",
                "seoScore": 0 if no_site else 45,
                "opportunityScore": 94 if no_site else 65,
                "recommendedPackage": "Website Launch + Local SEO" if no_site else "SEO Audit + Web Design Review",
            })
        return {
            "ok": True, "configured": True, "mode": "api", "provider": "google-places",
            "providerStatus": status, "items": items,
            "disclaimer": "Google Places results are public business candidates, not medical-quality rankings. Verify identity, license, official website and contact details independently.",
        }

    brave_key = os.getenv("BRAVE_SEARCH_API_KEY", "")
    if brave_key:
        params = urlencode({"q": combined, "count": 20, "search_lang": "fa", "safesearch": "strict"})
        status, response = get_json("https://api.search.brave.com/res/v1/web/search?" + params,
                                    {"X-Subscription-Token": brave_key})
        results = ((response.get("web") or {}).get("results") or []) if isinstance(response, dict) else []
        items = []
        for result in results[:20]:
            if not isinstance(result, dict):
                continue
            items.append({"name": str(result.get("title", ""))[:180],
                          "website": str(result.get("url", ""))[:500],
                          "phone": "", "email": "", "whatsapp": "", "whatsappLinks": [],
                          "address": "", "specialty": specialty, "tags": [],
                          "source": str(result.get("url", ""))[:500],
                          "summary": str(result.get("description", ""))[:600],
                          "verified": False})
        return {"ok": True, "configured": True, "mode": "api", "provider": "brave", "providerStatus": status,
                "items": items,
                "disclaimer": "Brave web results are discovery candidates, not verified medical providers. Confirm license, identity and public contact information."}
    encoded = quote_plus(combined)
    directory_query = quote_plus(f"site:paziresh24.com OR site:nobat.ir {combined}")
    return {
        "ok": True,
        "configured": False,
        "mode": "links",
        "items": [],
        "requiredConfiguration": ["GOOGLE_PLACES_API_KEY", "BRAVE_SEARCH_API_KEY", "CLINIC_SEARCH_WEBHOOK_URL"],
        "searchLinks": {
            "duckduckgo": "https://duckduckgo.com/?q=" + encoded,
            "google": "https://www.google.com/search?q=" + encoded,
            "bing": "https://www.bing.com/search?q=" + encoded,
            "brave": "https://search.brave.com/search?q=" + encoded,
            "medicalDirectories": "https://www.google.com/search?q=" + directory_query,
        },
        "disclaimer": "No clinic-search adapter is configured. Links search public business information only; do not collect patient data or infer sensitive traits.",
    }


def normalize_search_result_url(href: str, base_url: str = ""):
    if not href:
        return ""
    absolute = urljoin(base_url, href)
    parsed = urlparse(absolute)
    query = parse_qs(parsed.query)
    if parsed.path == "/url" and query.get("q"):
        absolute = query["q"][0]
    elif query.get("uddg"):
        absolute = unquote(query["uddg"][0])
    parsed = urlparse(absolute)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        return ""
    blocked = {"google.com", "www.google.com", "bing.com", "www.bing.com",
               "duckduckgo.com", "www.duckduckgo.com", "search.brave.com"}
    if parsed.netloc.lower() in blocked:
        return ""
    return absolute.split("#", 1)[0]


def clean_search_result_title(anchor):
    heading = anchor.find(["h1", "h2", "h3"]) if hasattr(anchor, "find") else None
    title = " ".join((heading or anchor).get_text(" ", strip=True).split())
    title = re.sub(r"https?://\S+", "", title, flags=re.IGNORECASE)
    title = re.sub(r"\b(?:www\.)?[a-zA-Z0-9][a-zA-Z0-9.-]+\.[a-zA-Z]{2,}\b", "", title, flags=re.IGNORECASE)
    title = re.sub(r"\s*[›»|]\s*.*$", "", title)
    title = re.sub(r"\s{2,}", " ", title).strip(" -–—|›")
    return title


def classify_search_candidate(title: str, url: str):
    combined = f"{title} {url}".lower()
    list_terms = ("بهترین", "لیست", "معرفی", "10 ", "۱۰ ", "راهنما", "قیمت", "هزینه", "best-", "top-", "list-")
    directory_terms = ("profile", "clinicdetail", "jobsearch", "/doctor", "/dr/", "directory")
    medical_terms = ("کلینیک", "درمانگاه", "مرکز پزشکی", "دکتر", "پزشک", "clinic", "medical", "dermatology", "dental")
    if any(term in combined for term in list_terms):
        return "list-article", "Extract clinics from article before adding as leads"
    if any(term in combined for term in directory_terms):
        return "directory-profile", "Verify profile identity and resolve official website"
    if any(term in combined for term in medical_terms):
        return "clinic-candidate", "Verify license, address and public contact details"
    return "web-result", "Review before adding as a clinic lead"


def parse_search_html(payload: dict):
    html = str(payload.get("html", ""))
    if not html.strip():
        raise ValueError("HTML content is required.")
    if len(html.encode("utf-8")) > 1_500_000:
        raise ValueError("HTML input is larger than 1.5 MB.")
    engine = str(payload.get("engine", "generic")).lower()[:30]
    source_url = str(payload.get("sourceUrl", ""))[:500]
    specialty = str(payload.get("specialty", "medical clinic"))[:150]
    soup = BeautifulSoup(html, "html.parser")
    items = []
    seen = set()

    selector_map = {
        "google": ["a:has(h3)", "div.MjjYud a[href]"],
        "bing": ["li.b_algo h2 a", "h2 a[href]"],
        "duckduckgo": ["a.result__a", ".result__title a"],
        "brave": ["a.result-header", "a[href]:has(h3)"],
        "generic": ["h2 a[href]", "h3 a[href]", "a[href]"],
    }
    selectors = selector_map.get(engine, selector_map["generic"])
    anchors = []
    for selector in selectors:
        try:
            anchors.extend(soup.select(selector))
        except Exception:
            continue
    for anchor in anchors:
        if anchor.name != "a":
            anchor = anchor.find("a", href=True)
        if not anchor or not anchor.get("href"):
            continue
        url = normalize_search_result_url(anchor.get("href", ""), source_url)
        if not url or url in seen:
            continue
        title = clean_search_result_title(anchor)
        if len(title) < 3:
            continue
        seen.add(url)
        container = anchor.find_parent(["article", "li", "div"]) or anchor.parent
        text = " ".join(container.get_text(" ", strip=True).split()) if container else title
        text = re.sub(r"https?://\S+", "", text)
        text = re.sub(r"\s{2,}", " ", text).strip()
        phone_match = re.search(r"(?:\+?98|0)?(?:21[-\s]?\d{5,8}|9\d{9})", text)
        signals = extract_public_contact_signals(str(container or anchor), source_url or url)
        result_type, action = classify_search_candidate(title, url)
        items.append({"name": title[:180], "website": url[:500],
                      "domain": (urlparse(url).hostname or "")[:200],
                      "phone": (signals["phoneNumbers"][0] if signals["phoneNumbers"] else phone_match.group(0) if phone_match else ""),
                      "email": signals["emails"][0] if signals["emails"] else "",
                      "whatsapp": signals["whatsappNumber"], "whatsappLinks": signals["whatsappLinks"],
                      "tags": signals["tags"], "address": signals["addresses"][0] if signals["addresses"] else "",
                      "specialty": specialty, "source": source_url or url[:500],
                      "summary": text[:600], "resultType": result_type,
                      "recommendedAction": action, "verified": False})
        if len(items) >= 50:
            break

    # Add structured medical entities from directory pages when available.
    for script in soup.find_all("script", attrs={"type": "application/ld+json"}):
        try:
            data = json.loads(script.get_text(strip=True) or "{}")
        except Exception:
            continue
        queue = data if isinstance(data, list) else [data]
        while queue:
            node = queue.pop(0)
            if isinstance(node, list):
                queue.extend(node)
                continue
            if not isinstance(node, dict):
                continue
            queue.extend(v for v in node.values() if isinstance(v, (dict, list)))
            node_type = node.get("@type", "")
            types = node_type if isinstance(node_type, list) else [node_type]
            if not any(str(t) in {"MedicalClinic", "MedicalBusiness", "Physician", "Dentist", "LocalBusiness"} for t in types):
                continue
            name = str(node.get("name", "")).strip()
            url = normalize_search_result_url(str(node.get("url", "")), source_url)
            key = url or name
            if not name or key in seen:
                continue
            address = node.get("address", "")
            if isinstance(address, dict):
                address = "، ".join(str(address.get(k, "")) for k in ("addressLocality", "streetAddress") if address.get(k))
            seen.add(key)
            same_as = node.get("sameAs", [])
            if not isinstance(same_as, list):
                same_as = [same_as]
            whatsapp_links = [str(x)[:500] for x in same_as if any(h in str(x).lower() for h in WHATSAPP_HOSTS)]
            whatsapp_number = ""
            if whatsapp_links:
                whatsapp_number = extract_public_contact_signals(f'<a href="{whatsapp_links[0]}">WhatsApp</a>', url or source_url)["whatsappNumber"]
            node_tags = node.get("keywords", [])
            if not isinstance(node_tags, list):
                node_tags = [x.strip() for x in re.split(r"[,،|]", str(node_tags)) if x.strip()]
            items.append({"name": name[:180], "website": url[:500],
                          "domain": (urlparse(url).hostname or "")[:200],
                          "phone": normalize_public_phone(str(node.get("telephone", "")))[:100],
                          "email": str(node.get("email", "")).removeprefix("mailto:")[:180],
                          "whatsapp": whatsapp_number, "whatsappLinks": whatsapp_links,
                          "tags": [str(x)[:80] for x in node_tags[:20]], "address": str(address)[:500],
                          "specialty": specialty, "source": source_url or url,
                          "summary": str(node.get("description", ""))[:600],
                          "resultType": "structured-medical-entity",
                          "recommendedAction": "Verify license and official ownership", "verified": False})
            if len(items) >= 50:
                break
    return {"ok": True, "engine": engine, "items": items,
            "count": len(items),
            "disclaimer": "Imported search results are unverified candidates. Confirm identity, medical license and active public contact details."}


def enrich_clinic_candidates(payload: dict):
    candidates = payload.get("items") if isinstance(payload.get("items"), list) else []
    specialty = str(payload.get("specialty", "medical clinic"))[:150]
    if not candidates:
        raise ValueError("At least one candidate is required for enrichment.")
    enriched, errors, seen = [], [], set()
    for candidate in candidates[:6]:
        if not isinstance(candidate, dict):
            continue
        url = str(candidate.get("website", "")).strip()
        if not url:
            continue
        ok, reason = public_url(url)
        if not ok:
            errors.append({"url": url[:500], "error": reason})
            continue
        if os.getenv("ENRICH_REQUIRE_ROBOTS", "false").lower() == "true" and not robots_allows(url):
            errors.append({"url": url[:500], "error": "robots.txt did not allow enrichment"})
            continue
        try:
            status, final_url, html, content_type, _ = fetch(url, timeout=15, limit=1_500_000)
            if status != 200 or ("html" not in content_type.lower() and "<html" not in html[:1000].lower()):
                raise ValueError(f"HTTP {status} or non-HTML response")
            contact_signals = extract_public_contact_signals(html, final_url)
            parsed = parse_search_html({"html": html, "engine": "generic", "sourceUrl": final_url,
                                        "specialty": specialty}).get("items", [])
            useful = []
            for item in parsed:
                kind = item.get("resultType")
                if kind in {"clinic-candidate", "directory-profile", "structured-medical-entity"}:
                    useful.append(item)
            if useful:
                for item in useful[:15]:
                    key = item.get("website") or item.get("name")
                    if not key or key in seen:
                        continue
                    seen.add(key)
                    item["parentSource"] = final_url
                    item["enriched"] = True
                    enriched.append(item)
            else:
                parser = AuditParser(final_url)
                parser.feed(html)
                name = str(candidate.get("name", "")).strip() or parser.title or (urlparse(final_url).hostname or "Clinic")
                kind, action = classify_search_candidate(name, final_url)
                key = final_url
                if key not in seen:
                    seen.add(key)
                    enriched.append({"name": name[:180], "website": final_url[:500],
                                     "domain": (urlparse(final_url).hostname or "")[:200],
                                     "phone": (contact_signals["phoneNumbers"][0] if contact_signals["phoneNumbers"] else str(candidate.get("phone", "")))[:100],
                                     "email": (contact_signals["emails"][0] if contact_signals["emails"] else str(candidate.get("email", "")))[:180],
                                     "whatsapp": contact_signals["whatsappNumber"] or str(candidate.get("whatsapp", ""))[:100],
                                     "whatsappLinks": contact_signals["whatsappLinks"], "tags": contact_signals["tags"],
                                     "address": (contact_signals["addresses"][0] if contact_signals["addresses"] else str(candidate.get("address", "")))[:500],
                                     "specialty": specialty, "source": final_url,
                                     "summary": parser.description[:600], "resultType": kind,
                                     "recommendedAction": action, "verified": False, "enriched": True})
        except Exception as exc:
            errors.append({"url": url[:500], "error": str(exc)[:300]})
    return {"ok": True, "items": enriched[:60], "count": len(enriched[:60]), "errors": errors,
            "disclaimer": "Enrichment resolves public pages into unverified clinic candidates. Verify license, identity and public contact details before outreach."}


def scraper_allowed_domains():
    return {item.strip().lower().lstrip(".") for item in
            re.split(r"[,\n]", os.getenv("SCRAPER_ALLOWED_DOMAINS", "")) if item.strip()}


def domain_is_allowed(host: str, allowed: set[str]):
    host = host.lower().rstrip(".")
    return any(host == domain or host.endswith("." + domain) for domain in allowed)


def robots_allows(url: str):
    if os.getenv("SCRAPER_IGNORE_ROBOTS", "false").lower() == "true":
        return True
    parsed = urlparse(url)
    robots_url = f"{parsed.scheme}://{parsed.netloc}/robots.txt"
    try:
        status, _, text, _, _ = fetch(robots_url, timeout=8, limit=300_000)
        if status != 200:
            return False
        parser = RobotFileParser()
        parser.set_url(robots_url)
        parser.parse(text.splitlines())
        return parser.can_fetch(USER_AGENT, url)
    except Exception:
        return False


def scrape_clinic_directory(payload: dict):
    url = str(payload.get("url", "")).strip()
    specialty = str(payload.get("specialty", "medical clinic"))[:150]
    if not url:
        raise ValueError("Directory URL is required.")
    ok, reason = public_url(url)
    if not ok:
        raise ValueError(reason)
    allowed = scraper_allowed_domains()
    host = urlparse(url).hostname or ""
    if not allowed:
        raise ValueError("SCRAPER_ALLOWED_DOMAINS is empty. Add approved public directory domains before server-side scraping.")
    if not domain_is_allowed(host, allowed):
        raise ValueError("This domain is not in SCRAPER_ALLOWED_DOMAINS.")
    forbidden = {"google.com", "bing.com", "duckduckgo.com", "search.brave.com"}
    if any(host == item or host.endswith("." + item) for item in forbidden):
        raise ValueError("Automatic scraping of search-engine result pages is disabled. Use HTML import or an approved search API.")
    if not robots_allows(url):
        raise ValueError("robots.txt does not allow this scraper or could not be verified.")
    status, final_url, html, content_type, elapsed = fetch(url, timeout=20, limit=1_500_000)
    if status != 200 or ("html" not in content_type.lower() and "<html" not in html[:1000].lower()):
        raise ValueError(f"Directory returned HTTP {status} or non-HTML content.")
    result = parse_search_html({"html": html, "engine": "generic", "sourceUrl": final_url,
                                "specialty": specialty})
    result.update({"url": final_url, "elapsedSeconds": elapsed, "robotsAllowed": True})
    return result


def run_configured_discovery():
    raw_urls = os.getenv("CLINIC_DISCOVERY_URLS", "")
    urls = [item.strip() for item in re.split(r"[,\n]", raw_urls) if item.strip()][:5]
    if not urls:
        return {"ok": True, "skipped": True, "message": "CLINIC_DISCOVERY_URLS is empty.", "items": []}
    all_items, errors = [], []
    seen = set()
    for url in urls:
        try:
            result = scrape_clinic_directory({"url": url, "specialty": "medical clinic"})
            for item in result.get("items", []):
                key = item.get("website") or item.get("name")
                if key and key not in seen:
                    seen.add(key)
                    all_items.append(item)
        except Exception as exc:
            errors.append({"url": url, "error": str(exc)[:300]})
    delivered = False
    persistence = None
    if all_items and (os.getenv("SUPABASE_URL") or os.getenv("LEAD_DATABASE_WEBHOOK_URL") or os.getenv("LEAD_INGEST_WEBHOOK_URL")):
        try:
            persistence = persist_leads_database(all_items)
            delivered = True
        except Exception as exc:
            errors.append({"url": "database", "error": str(exc)[:300]})
    return {"ok": True, "skipped": False, "count": len(all_items), "items": all_items[:100],
            "errors": errors, "persisted": delivered, "persistence": persistence,
            "warning": "Without Supabase or a lead database webhook, serverless cron results are not persisted."}


def clinic_export_rows(items: list[dict]):
    rows = []
    for item in items[:500]:
        if not isinstance(item, dict):
            continue
        ai = item.get("aiAnalysis") if isinstance(item.get("aiAnalysis"), dict) else {}
        rows.append({
            "name": str(ai.get("normalizedName") or item.get("name", ""))[:250],
            "website": str(item.get("website", ""))[:500],
            "phone": str(item.get("phone", ""))[:100],
            "address": str(item.get("address", ""))[:500],
            "specialty": str(ai.get("specialty") or item.get("specialty", ""))[:180],
            "result_type": str(ai.get("resultType") or item.get("resultType", "candidate"))[:80],
            "ai_confidence": ai.get("confidence", ""),
            "ai_priority": str(ai.get("priority", ""))[:30],
            "ai_reason": str(ai.get("reason", ""))[:500],
            "recommended_next_step": str(ai.get("recommendedNextStep") or item.get("recommendedAction", ""))[:500],
            "source": str(item.get("source", ""))[:500],
            "verified": bool(item.get("verified", False)),
        })
    return rows


def export_clinic_candidates(payload: dict):
    items = payload.get("items") if isinstance(payload.get("items"), list) else []
    rows = clinic_export_rows(items)
    if not rows:
        raise ValueError("No clinic candidates were supplied for export.")
    fmt = str(payload.get("format", "csv")).lower()
    title = str(payload.get("title", "Clinic Discovery Results"))[:150]
    columns = ["name", "website", "phone", "address", "specialty", "result_type", "ai_confidence", "ai_priority", "ai_reason", "recommended_next_step", "source", "verified"]
    if fmt == "csv":
        stream = StringIO()
        writer = csv.DictWriter(stream, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)
        return ("\ufeff" + stream.getvalue()).encode("utf-8"), "clinic-discovery.csv", "text/csv; charset=utf-8"
    if fmt == "xlsx":
        if not OPENPYXL_AVAILABLE:
            raise ValueError("Excel export requires openpyxl.")
        wb = Workbook()
        ws = wb.active
        ws.title = "Clinic Leads"
        ws.sheet_view.rightToLeft = True
        ws.freeze_panes = "A2"
        header_fill = PatternFill("solid", fgColor="17324D")
        for col, name in enumerate(columns, 1):
            cell = ws.cell(1, col, name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row_index, row in enumerate(rows, 2):
            for col, name in enumerate(columns, 1):
                cell = ws.cell(row_index, col, row.get(name, ""))
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        widths = [28, 42, 18, 38, 25, 22, 14, 14, 45, 45, 42, 12]
        for index, width in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + index)].width = width
        output = BytesIO()
        wb.save(output)
        return output.getvalue(), "clinic-discovery.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if fmt == "pdf":
        if not PILLOW_AVAILABLE:
            raise ValueError("PDF export requires Pillow.")
        width, height = 1654, 1169
        margin = 65
        bundled = ROOT / "assets" / "fonts"
        regular_path = str(bundled / "DejaVuSans.ttf")
        bold_path = str(bundled / "DejaVuSans-Bold.ttf")
        layout = ImageFont.Layout.RAQM if hasattr(ImageFont, "Layout") and RAQM_AVAILABLE else ImageFont.Layout.BASIC if hasattr(ImageFont, "Layout") else None
        regular = ImageFont.truetype(regular_path, 20, layout_engine=layout)
        small = ImageFont.truetype(regular_path, 15, layout_engine=layout)
        bold_font = ImageFont.truetype(bold_path, 27, layout_engine=layout)
        row_font = ImageFont.truetype(bold_path, 18, layout_engine=layout)
        def shape(value):
            value = str(value)
            if RAQM_AVAILABLE or not re.search(r"[\u0600-\u06FF]", value):
                return value
            if BIDI_FALLBACK_AVAILABLE:
                try: return bidi_get_display(arabic_reshaper.reshape(value))
                except Exception: return value
            return value
        def draw_text(draw, xy, value, font, fill, anchor="ra"):
            if RAQM_AVAILABLE:
                try:
                    draw.text(xy, str(value), font=font, fill=fill, anchor=anchor, direction="rtl", language="fa")
                    return
                except (ValueError, TypeError, KeyError): pass
            draw.text(xy, shape(value), font=font, fill=fill, anchor=anchor)
        pages = []
        per_page = 10
        for offset in range(0, len(rows), per_page):
            page = Image.new("RGB", (width, height), "white")
            draw = ImageDraw.Draw(page)
            draw.rectangle((0, 0, width, 105), fill="#17324D")
            draw_text(draw, (width-margin, 34), title, bold_font, "white")
            draw_text(draw, (margin, 42), f"{offset+1}-{min(offset+per_page,len(rows))} / {len(rows)}", regular, "#C6D7E4", anchor="la")
            y = 130
            for number, row in enumerate(rows[offset:offset+per_page], offset+1):
                draw.rounded_rectangle((margin, y, width-margin, y+86), radius=12, fill="#F4F7F9", outline="#DCE6EE")
                draw_text(draw, (width-margin-18, y+12), f"{number}. {row['name']}", row_font, "#17324D")
                details = f"{row['specialty']} | {row['result_type']} | AI: {row['ai_confidence'] or '—'} | {row['phone'] or '—'}"
                draw_text(draw, (width-margin-18, y+43), details[:150], small, "#526B7C")
                draw.text((margin+18, y+60), row['website'][:110], font=small, fill="#246BFD", anchor="ls")
                y += 94
            pages.append(page)
        output = BytesIO()
        pages[0].save(output, format="PDF", save_all=True, append_images=pages[1:], resolution=150.0, title=title)
        return output.getvalue(), "clinic-discovery.pdf", "application/pdf"
    raise ValueError("Export format must be csv, xlsx or pdf.")


def first_environment_value(*names):
    for name in names:
        value = os.getenv(name, "").strip()
        if value:
            return value, name
    return "", ""


def supabase_settings():
    url, url_name = first_environment_value("SUPABASE_URL", "NEXT_PUBLIC_SUPABASE_URL", "VITE_SUPABASE_URL", "PUBLIC_SUPABASE_URL")
    key, key_name = first_environment_value("SUPABASE_SERVICE_ROLE_KEY", "SUPABASE_SECRET_KEY", "SUPABASE_SERVICE_KEY")
    table = re.sub(r"[^a-zA-Z0-9_]", "", os.getenv("SUPABASE_LEADS_TABLE", "clinic_leads")) or "clinic_leads"
    return {"url": url.rstrip("/"), "key": key, "table": table, "urlVariable": url_name, "keyVariable": key_name,
            "configured": bool(url and key)}


def lead_dedupe_key(item: dict, website: str) -> str:
    if website:
        try:
            parsed = urlparse(website if urlparse(website).scheme else "https://" + website)
            host = (parsed.hostname or "").lower().removeprefix("www.")
            if host:
                return "website:" + host[:220]
        except Exception:
            pass
    identity = "|".join([
        str(item.get("placeId", "")), str(item.get("name", "")),
        str(item.get("phone", "")), str(item.get("address", item.get("area", ""))),
        str(item.get("source", "")),
    ]).strip("|").lower()
    return "entity:" + hashlib.sha256(identity.encode("utf-8")).hexdigest()


def normalize_lead_row(item: dict):
    website = str(item.get("website", "")).strip()[:500]
    return {
        "dedupe_key": lead_dedupe_key(item, website),
        "name": str(item.get("name", "Clinic candidate"))[:180],
        "website": website,
        "phone": str(item.get("phone", ""))[:100],
        "email": str(item.get("email", ""))[:180],
        "whatsapp": str(item.get("whatsapp", item.get("whatsappNumber", "")))[:100],
        "tags": [str(x)[:80] for x in item.get("tags", [])[:40]] if isinstance(item.get("tags"), list) else [],
        "address": str(item.get("address", item.get("area", "")))[:500],
        "specialty": str(item.get("specialty", item.get("services", "")))[:180],
        "source": str(item.get("source", ""))[:500],
        "result_type": str(item.get("resultType", "candidate"))[:80],
        "status": str(item.get("status", "new"))[:50],
        "seo_score": int(item.get("seo", item.get("seoScore", 0)) or 0),
        "opportunity_score": int(item.get("opportunity", item.get("opportunityScore", 0)) or 0),
        "raw": item,
    }


def persist_leads_database(items: list[dict]):
    rows = [normalize_lead_row(item) for item in items[:100] if isinstance(item, dict)]
    rows = [row for row in rows if row["name"] and (row["website"] or row["phone"] or row["email"] or row["whatsapp"] or row["address"] or row["source"])]
    if not rows:
        raise ValueError("No valid lead rows were supplied. A name plus website, phone, address or source is required.")
    settings = supabase_settings()
    supabase_url, supabase_key, table = settings["url"], settings["key"], settings["table"]
    if settings["configured"]:
        endpoint = f"{supabase_url}/rest/v1/{table}?on_conflict=dedupe_key"
        response = requests.post(endpoint, headers={"apikey": supabase_key,
            "Authorization": f"Bearer {supabase_key}", "Content-Type": "application/json",
            "Prefer": "resolution=merge-duplicates,return=representation"}, json=rows, timeout=30)
        if not response.ok:
            raise ValueError(f"Supabase HTTP {response.status_code}: {response.text[:500]}")
        data = response.json() if response.text else []
        return {"ok": True, "provider": "supabase", "saved": len(rows), "items": data,
                "detectedVariables": {"url": settings["urlVariable"], "key": settings["keyVariable"]},
                "table": table}
    webhook = os.getenv("LEAD_DATABASE_WEBHOOK_URL", "") or os.getenv("LEAD_INGEST_WEBHOOK_URL", "")
    if webhook:
        token = os.getenv("LEAD_DATABASE_WEBHOOK_TOKEN", "") or os.getenv("LEAD_INGEST_WEBHOOK_TOKEN", "")
        headers = {"Authorization": f"Bearer {token}"} if token else {}
        status, data = post_json(webhook, {"source": "clinic-signal", "items": rows}, headers)
        return {"ok": True, "provider": "webhook", "providerStatus": status, "saved": len(rows), "response": data}
    raise ValueError("No lead database is configured. Add Supabase variables or LEAD_DATABASE_WEBHOOK_URL.")


def fetch_leads_database(limit: int = 100):
    settings = supabase_settings()
    supabase_url, supabase_key, table = settings["url"], settings["key"], settings["table"]
    if not settings["configured"]:
        raise ValueError("Supabase lead database is not configured. Expected URL: SUPABASE_URL or NEXT_PUBLIC_SUPABASE_URL; expected server key: SUPABASE_SERVICE_ROLE_KEY or SUPABASE_SECRET_KEY.")
    endpoint = f"{supabase_url}/rest/v1/{table}?select=*&order=created_at.desc&limit={max(1,min(limit,500))}"
    response = requests.get(endpoint, headers={"apikey": supabase_key,
        "Authorization": f"Bearer {supabase_key}"}, timeout=30)
    if not response.ok:
        raise ValueError(f"Supabase HTTP {response.status_code}: {response.text[:500]}")
    return {"ok": True, "provider": "supabase", "items": response.json(), "table": table,
            "detectedVariables": {"url": settings["urlVariable"], "key": settings["keyVariable"]}}


EXHIBITION_FIELD_ALIASES = {
    "company": {"company", "company name", "exhibitor", "name", "شرکت", "نام شرکت", "مشارکت کننده", "غرفه دار"},
    "booth": {"booth", "stand", "hall/booth", "غرفه", "شماره غرفه", "سالن و غرفه"},
    "category": {"category", "industry", "sector", "محصول", "گروه", "حوزه فعالیت", "صنعت"},
    "phone": {"phone", "telephone", "mobile", "tel", "تلفن", "شماره تماس", "موبایل"},
    "website": {"website", "site", "url", "وب سایت", "وب‌سایت", "سایت"},
    "email": {"email", "e-mail", "ایمیل", "پست الکترونیک"},
    "city": {"city", "location", "شهر", "استان", "موقعیت"},
}


def map_exhibition_header(value: str):
    normalized = re.sub(r"\s+", " ", str(value).strip().lower())
    for field, aliases in EXHIBITION_FIELD_ALIASES.items():
        if normalized in aliases:
            return field
    return normalized


def normalize_exhibitor(row: dict, event: dict):
    mapped = {}
    for key, value in row.items():
        mapped[map_exhibition_header(key)] = " ".join(str(value or "").split())
    company = mapped.get("company") or mapped.get("نام") or next((v for v in mapped.values() if v), "")
    website = mapped.get("website", "").strip()
    if website and not website.startswith(("http://", "https://")):
        website = "https://" + website
    return {"name": company[:220], "website": website[:500], "phone": mapped.get("phone", "")[:100],
            "email": mapped.get("email", "")[:180], "city": mapped.get("city", "")[:150],
            "booth": mapped.get("booth", "")[:100], "category": mapped.get("category", "")[:180],
            "eventName": str(event.get("name", ""))[:220], "eventDate": str(event.get("date", ""))[:100],
            "eventLocation": str(event.get("location", ""))[:220], "eventSource": str(event.get("source", ""))[:500],
            "source": "exhibition-import", "resultType": "exhibitor", "verified": False, "raw": mapped}


def parse_exhibition_data(payload: dict):
    raw = str(payload.get("data", ""))
    if not raw.strip():
        raise ValueError("Exhibition list data is required.")
    if len(raw.encode("utf-8")) > 2_000_000:
        raise ValueError("Exhibition import is larger than 2 MB.")
    event = payload.get("event") if isinstance(payload.get("event"), dict) else {}
    fmt = str(payload.get("format", "auto")).lower()
    items = []
    if fmt == "html" or (fmt == "auto" and re.search(r"<table|<tr|<td", raw, re.IGNORECASE)):
        soup = BeautifulSoup(raw, "html.parser")
        for table in soup.find_all("table"):
            rows = table.find_all("tr")
            if not rows:
                continue
            header_cells = rows[0].find_all(["th", "td"])
            headers = [cell.get_text(" ", strip=True) or f"column_{i+1}" for i, cell in enumerate(header_cells)]
            for tr in rows[1:]:
                cells = [cell.get_text(" ", strip=True) for cell in tr.find_all(["td", "th"])]
                if not cells:
                    continue
                row = {headers[i] if i < len(headers) else f"column_{i+1}": value for i, value in enumerate(cells)}
                item = normalize_exhibitor(row, event)
                if item["name"]:
                    items.append(item)
    else:
        sample = raw[:5000]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            has_header = csv.Sniffer().has_header(sample)
        except csv.Error:
            dialect, has_header = csv.excel, False
        stream = StringIO(raw)
        if has_header:
            reader = csv.DictReader(stream, dialect=dialect)
            for row in reader:
                item = normalize_exhibitor(dict(row), event)
                if item["name"]:
                    items.append(item)
        else:
            reader = csv.reader(stream, dialect=dialect)
            for cells in reader:
                cells = [" ".join(str(cell).split()) for cell in cells]
                if not any(cells):
                    continue
                row = {"company": cells[0]}
                if len(cells) > 1: row["booth"] = cells[1]
                if len(cells) > 2: row["category"] = cells[2]
                if len(cells) > 3: row["phone"] = cells[3]
                if len(cells) > 4: row["website"] = cells[4]
                item = normalize_exhibitor(row, event)
                if item["name"]:
                    items.append(item)
    deduped, seen = [], set()
    for item in items[:1000]:
        key = re.sub(r"\W+", "", item["name"].lower())
        if key and key not in seen:
            seen.add(key)
            deduped.append(item)
    return {"ok": True, "items": deduped, "count": len(deduped), "event": event,
            "disclaimer": "Exhibitor data is imported as unverified public-business leads. Confirm identity and contact details before outreach."}


def find_company_website_brave(company: str, city: str = ""):
    key = os.getenv("BRAVE_SEARCH_API_KEY", "")
    query = f'"{company}" {city} official website وب سایت رسمی'.strip()
    links = {"google": "https://www.google.com/search?q=" + quote_plus(query),
             "duckduckgo": "https://duckduckgo.com/?q=" + quote_plus(query),
             "bing": "https://www.bing.com/search?q=" + quote_plus(query)}
    if not key:
        return "", links, "links"
    params = urlencode({"q": query, "count": 10, "search_lang": "fa", "safesearch": "strict"})
    _, response = get_json("https://api.search.brave.com/res/v1/web/search?" + params,
                           {"X-Subscription-Token": key})
    results = ((response.get("web") or {}).get("results") or []) if isinstance(response, dict) else []
    blocked = ("instagram.com", "linkedin.com", "facebook.com", "t.me", "wikipedia.org", "paziresh24.com", "nobat.ir")
    for result in results:
        url = str(result.get("url", ""))
        host = (urlparse(url).hostname or "").lower()
        if url.startswith("http") and not any(host == b or host.endswith("." + b) for b in blocked):
            return url, links, "brave"
    return "", links, "brave"


def enrich_exhibition_companies(payload: dict):
    items = payload.get("items") if isinstance(payload.get("items"), list) else []
    run_audit = payload.get("audit") is not False
    if not items:
        raise ValueError("Select exhibition companies for enrichment.")
    output = []
    for item in items[:8]:
        if not isinstance(item, dict):
            continue
        enriched = dict(item)
        website = str(enriched.get("website", "")).strip()
        links = {}
        discovery_mode = "provided"
        if not website:
            website, links, discovery_mode = find_company_website_brave(str(enriched.get("name", "")), str(enriched.get("city", "")))
            enriched["website"] = website
        enriched["websiteSearchLinks"] = links
        enriched["websiteDiscoveryMode"] = discovery_mode
        if website and run_audit:
            try:
                report = audit(website)
                enriched["audit"] = report
                score = int(report.get("seoScore", 0) or 0)
                enriched["seoScore"] = score
                enriched["websiteStatus"] = "working" if report.get("status") == 200 else "error"
                enriched["opportunityScore"] = max(20, min(95, round((100-score)*0.75 + 25)))
                enriched["recommendedPackage"] = "Technical SEO Recovery" if score < 50 else "SEO Growth 90 Days" if score < 80 else "Content & CRO Growth"
            except Exception as exc:
                enriched["websiteStatus"] = "audit-error"
                enriched["auditError"] = str(exc)[:300]
                enriched["opportunityScore"] = 82
                enriched["recommendedPackage"] = "Website Technical Recovery"
        elif not website:
            enriched["websiteStatus"] = "no-website-found"
            enriched["seoScore"] = 0
            enriched["opportunityScore"] = 94
            enriched["recommendedPackage"] = "Website Launch + Local SEO"
        else:
            enriched["websiteStatus"] = "website-found"
            enriched["opportunityScore"] = 60
            enriched["recommendedPackage"] = "Website Audit Required"
        output.append(enriched)
    return {"ok": True, "items": output, "count": len(output),
            "searchProviderConfigured": bool(os.getenv("BRAVE_SEARCH_API_KEY")),
            "disclaimer": "Website matches and opportunity scores are unverified sales research. Confirm official ownership before outreach."}


def make_proposal_pdf(payload: dict) -> tuple[bytes, str]:
    """Create a real A4 PDF with Pillow/RAQM; browser print remains the fallback."""
    if not PILLOW_AVAILABLE:
        raise ValueError("Direct PDF rendering is unavailable; use browser Print / Save as PDF.")
    lead = payload.get("lead") if isinstance(payload.get("lead"), dict) else {}
    def val(source, key, default="", limit=1500):
        return str(source.get(key, default))[:limit]
    name = val(lead, "name", "Clinic")
    agency_profile = payload.get("agencyProfile") if isinstance(payload.get("agencyProfile"), dict) else {}
    agency = val(agency_profile, "name", val(payload, "agency", "Clinic Signal Partner", 160), 160)
    agency_phone = val(agency_profile, "phone", "", 80)
    agency_website = val(agency_profile, "website", "", 300)
    agency_email = val(agency_profile, "email", "", 200)
    agency_address = val(agency_profile, "address", "", 500)
    agency_hours = val(agency_profile, "hours", "", 200)
    logo_data = val(agency_profile, "logoData", "", 900_000)
    issue = val(lead, "issue", "Technical and organic growth opportunity")
    tech = val(lead, "tech", "Public technical audit pending")
    plan = val(lead, "plan", "Technical remediation, local landing pages and conversion tracking")
    target = val(lead, "target", "Service + location search clusters")
    package = val(lead, "package", "Growth package", 160)
    priority = val(lead, "priority", "P2", 10)
    validity = val(payload, "validity", "14 days", 80)
    setup = val(payload, "setup", "—", 120)
    monthly = val(payload, "monthly", "—", 120)
    media = val(payload, "media", "—", 120)
    duration = val(payload, "duration", "—", 120)
    try:
        seo = max(0, min(100, int(lead.get("seo", 0))))
        opportunity = max(0, min(100, int(lead.get("opportunity", 0))))
    except Exception:
        seo = opportunity = 0

    W, H = 1240, 1754
    image = Image.new("RGB", (W, H), "white")
    draw = ImageDraw.Draw(image)
    bundled_fonts = ROOT / "assets" / "fonts"
    regular_path = str(bundled_fonts / "DejaVuSans.ttf") if (bundled_fonts / "DejaVuSans.ttf").exists() else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
    bold_path = str(bundled_fonts / "DejaVuSans-Bold.ttf") if (bundled_fonts / "DejaVuSans-Bold.ttf").exists() else "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
    if hasattr(ImageFont, "Layout"):
        layout = ImageFont.Layout.RAQM if RAQM_AVAILABLE else ImageFont.Layout.BASIC
    else:
        layout = None
    regular = lambda n: ImageFont.truetype(regular_path, n, layout_engine=layout)
    bold = lambda n: ImageFont.truetype(bold_path, n, layout_engine=layout)
    f_small, f_body, f_bold, f_h2, f_title = regular(19), regular(23), bold(23), bold(30), bold(40)
    ink, muted, teal, pale, amber = "#183247", "#64748b", "#00a69c", "#f2f7f9", "#fff7e2"
    left, right = 82, W - 82

    def prepare_text(text):
        value = str(text)
        if RAQM_AVAILABLE or not re.search(r"[\u0600-\u06FF]", value):
            return value
        if BIDI_FALLBACK_AVAILABLE:
            try:
                return bidi_get_display(arabic_reshaper.reshape(value))
            except Exception:
                return value
        return value

    def text_rtl(x, y, text, font, fill=ink, anchor="ra"):
        kwargs = {"font": font, "fill": fill, "anchor": anchor}
        if RAQM_AVAILABLE:
            try:
                draw.text((x, y), str(text), direction="rtl", language="fa", **kwargs)
                return
            except (TypeError, ValueError, KeyError):
                pass
        draw.text((x, y), prepare_text(text), **kwargs)

    def measure(text, font):
        if RAQM_AVAILABLE:
            try:
                return draw.textlength(str(text), font=font, direction="rtl", language="fa")
            except (TypeError, ValueError, KeyError):
                pass
        return draw.textlength(prepare_text(text), font=font)

    def paragraph(text, y, font=f_body, fill=muted, max_width=None, line_height=37, max_lines=5):
        max_width = max_width or (right-left)
        words = str(text).split()
        lines, line = [], ""
        for word in words:
            candidate = (line + " " + word).strip()
            if measure(candidate, font) <= max_width:
                line = candidate
            else:
                if line:
                    lines.append(line)
                line = word
                if len(lines) >= max_lines:
                    break
        if line and len(lines) < max_lines:
            lines.append(line)
        if len(lines) == max_lines and len(words) > sum(len(x.split()) for x in lines):
            lines[-1] = lines[-1].rstrip("…") + "…"
        for line in lines:
            text_rtl(right, y, line, font, fill)
            y += line_height
        return y

    def heading(text, y):
        draw.rectangle((right-7, y-4, right, y+33), fill=teal)
        text_rtl(right-18, y, text, f_h2, ink)
        return y + 52

    # Header and sender logo
    logo_box = (right-92, 58, right, 142)
    logo_drawn = False
    if logo_data.startswith("data:image/") and ";base64," in logo_data:
        try:
            raw = base64.b64decode(logo_data.split(",", 1)[1], validate=True)
            logo_img = Image.open(BytesIO(raw)).convert("RGBA")
            logo_img.thumbnail((88, 78))
            px = right - 46 - logo_img.width//2
            py = 100 - logo_img.height//2
            image.paste(logo_img, (px, py), logo_img)
            logo_drawn = True
        except Exception:
            logo_drawn = False
    if not logo_drawn:
        draw.rounded_rectangle(logo_box, radius=16, fill="#12364d")
        text_rtl(right-46, 87, agency[:3], f_bold, "white", anchor="mm")
    text_rtl(right-112, 67, agency, f_h2, ink)
    text_rtl(right-112, 107, "پیشنهاد رشد ارگانیک و زیرساخت دیجیتال", f_small, muted)
    contact_line = " · ".join(x for x in (agency_phone, agency_website) if x)
    if contact_line:
        text_rtl(right-112, 137, contact_line, regular(15), muted)
    text_rtl(left, 78, time.strftime("%Y-%m-%d"), f_small, muted, anchor="la")
    text_rtl(left, 110, f"اعتبار: {validity}", f_small, muted, anchor="la")
    draw.rectangle((left, 158, right, 163), fill=teal)

    y = 190
    text_rtl(right, y, f"پیشنهاد اختصاصی برای {name}", f_title, ink)
    y += 68
    y = paragraph("این سند براساس بررسی عمومی حضور دیجیتال و وضعیت فنی مشاهده‌شده تهیه شده است. برآورد مقیاس به معنی درآمد واقعی یا توان پرداخت قطعی نیست.", y, f_body, muted, max_lines=3)
    y += 20

    # Score cards
    gap = 18
    box_w = (right-left-2*gap)//3
    cards = [("بلوغ SEO", f"{seo}/100"), ("فرصت", f"{opportunity}/100"), ("پکیج", package)]
    for i, (label, value) in enumerate(cards):
        x1 = right - (i+1)*box_w - i*gap
        x2 = x1 + box_w
        draw.rounded_rectangle((x1, y, x2, y+112), radius=16, fill=pale)
        text_rtl(x2-18, y+20, label, f_small, muted)
        text_rtl(x2-18, y+56, value, f_bold if i<2 else f_small, ink)
    text_rtl(left+16, y+17, priority, f_bold, "#a52d2d", anchor="la")
    y += 145

    y = heading("یافته و فرصت اصلی", y)
    y = paragraph("مشاهده فنی: " + tech, y, max_lines=3)
    y = paragraph("فرصت: " + issue, y+5, max_lines=3)
    y = paragraph("خوشه هدف: " + target, y+5, max_lines=2)
    y += 12

    y = heading("راهکار و برنامه ۹۰روزه", y)
    y = paragraph(plan, y, max_lines=3)
    phases = [
        ("روز ۱–۳۰", "Baseline، دسترسی‌ها و رفع ریسک فنی"),
        ("روز ۳۱–۶۰", "صفحات پول‌ساز، Schema و محتوای پزشکی"),
        ("روز ۶۱–۹۰", "CRO، Digital PR و گزارش لید"),
    ]
    phase_y = y + 10
    for i, (title, desc) in enumerate(phases):
        x1 = right - (i+1)*box_w - i*gap
        x2 = x1 + box_w
        draw.rounded_rectangle((x1, phase_y, x2, phase_y+105), radius=14, outline="#dce6ee", width=2)
        text_rtl(x2-14, phase_y+14, title, f_bold, ink)
        paragraph(desc, phase_y+50, f_small, muted, max_width=box_w-28, line_height=28, max_lines=2)
    y = phase_y + 130

    y = heading("سرمایه‌گذاری پیشنهادی", y)
    prices = [("راه‌اندازی", setup), ("حق‌الزحمه ماهانه", monthly), ("رسانه مستقیم", media), ("دوره", duration)]
    for label, value in prices:
        draw.line((left, y+35, right, y+35), fill="#dce6ee", width=1)
        text_rtl(right, y, label, f_body, muted)
        text_rtl(left, y, value, f_bold, ink, anchor="la")
        y += 43
    y += 12

    y = heading("KPI و شرایط حقوقی", y)
    y = paragraph("KPIها: دسترس‌پذیری، Core Web Vitals، رشد صفحات هدف، سهم Top 10، تماس و فرم واجدشرایط، نرخ تبدیل و در صورت اتصال CRM درآمد منتسب.", y, f_small, muted, line_height=31, max_lines=3)
    draw.rounded_rectangle((left, y+8, right, min(H-92, y+160)), radius=14, fill=amber, outline="#f0dfa8")
    paragraph("هیچ رتبه مطلق یا جایگاه ۱ تضمین نمی‌شود. تعهد مجری بر Deliverable، SLA، کیفیت فنی و KPIهای قابل‌اندازه‌گیری است. ادعاهای پزشکی فقط پس از تأیید پزشک مسئول منتشر می‌شود و جبران خدمت صرفاً طبق قرارداد خواهد بود.", y+24, f_small, "#725a20", max_width=right-left-34, line_height=30, max_lines=4)
    sender_footer = " · ".join(x for x in (agency, agency_phone, agency_email, agency_website) if x)
    address_footer = " · ".join(x for x in (agency_address, agency_hours) if x)
    if address_footer:
        text_rtl(right, H-112, address_footer, regular(13), muted)
    if sender_footer:
        text_rtl(right, H-86, sender_footer, regular(14), muted)
    text_rtl(right, H-58, "پیش‌نویس تجاری — نیازمند قرارداد و تأیید نهایی طرفین", regular(15), muted)

    output = BytesIO()
    image.save(output, format="PDF", resolution=150.0, title=f"Proposal for {name}", author=agency)
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "-", val(lead, "id", "clinic", 80)).strip("-") or "clinic"
    return output.getvalue(), f"proposal-{safe_name}.pdf"


def cleanup_pdf_links():
    now = time.time()
    for token in [k for k, v in PDF_LINKS.items() if v.get("expires", 0) <= now]:
        PDF_LINKS.pop(token, None)
    if len(PDF_LINKS) >= PDF_LINK_LIMIT:
        oldest = sorted(PDF_LINKS, key=lambda key: PDF_LINKS[key].get("created", 0))
        for token in oldest[:len(PDF_LINKS) - PDF_LINK_LIMIT + 1]:
            PDF_LINKS.pop(token, None)


def store_proposal_link(payload: dict):
    cleanup_pdf_links()
    pdf, filename = make_proposal_pdf(payload)
    token = secrets.token_urlsafe(24)
    now = time.time()
    PDF_LINKS[token] = {"pdf": pdf, "filename": filename, "created": now,
                        "expires": now + PDF_LINK_TTL, "downloads": 0}
    return token, filename, int(now + PDF_LINK_TTL)


class Handler(SimpleHTTPRequestHandler):
    server_version = "ClinicSignal/1.1"

    def translate_path(self, path):
        clean = urlparse(path).path.lstrip("/") or "app.html"
        target = (ROOT / clean).resolve()
        if ROOT not in target.parents and target != ROOT:
            return str(ROOT / "index.html")
        return str(target)

    def end_headers(self):
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("Referrer-Policy", "no-referrer")
        self.send_header("Content-Security-Policy", "default-src 'self' data: https:; script-src 'self' 'unsafe-inline'; style-src 'self' 'unsafe-inline'; img-src 'self' data: https:; connect-src 'self' https:; frame-ancestors 'self'")
        self.send_header("Cache-Control", "no-store")
        super().end_headers()

    def log_message(self, fmt, *args):
        print(f"[{self.log_date_time_string()}] {fmt % args}")

    def json_response(self, data, status=200):
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def public_base_url(self):
        configured = os.getenv("PUBLIC_BASE_URL", "").strip().rstrip("/")
        if configured:
            parsed = urlparse(configured)
            if parsed.scheme in {"http", "https"} and parsed.netloc:
                return configured
        proto = self.headers.get("X-Forwarded-Proto", "http").split(",", 1)[0].strip()
        if proto not in {"http", "https"}:
            proto = "http"
        host = self.headers.get("X-Forwarded-Host", self.headers.get("Host", "127.0.0.1:8000")).split(",", 1)[0].strip()
        if not re.fullmatch(r"[A-Za-z0-9.:[\]_-]+", host):
            host = "127.0.0.1:8000"
        return f"{proto}://{host}"

    def do_GET(self):
        path = urlparse(self.path).path
        auth_error = integration_auth_error(self.headers, path)
        if auth_error:
            status, message = auth_error
            return self.json_response({"ok": False, "error": message}, status)
        if path == "/api/health":
            verified = self.headers.get("X-Clinic-Signal-Internal", "") == "1"
            return self.json_response({"ok": True, "service": "Clinic Signal", "mode": "live-audit-and-messaging",
                                       "integrationAuth": "verified" if verified else "not-requested"})
        if path == "/api/integrations":
            return self.json_response(provider_status())
        if path == "/api/send-log":
            return self.json_response({"ok": True, "items": list(SEND_LOG)})
        if path == "/api/leads":
            try:
                return self.json_response(fetch_leads_database(int(parse_qs(urlparse(self.path).query).get("limit", ["100"])[0])))
            except Exception as exc:
                return self.json_response({"ok": False, "error": str(exc)}, 400)
        if path == "/api/run-discovery":
            secret = os.getenv("CRON_SECRET", "")
            if secret and self.headers.get("Authorization", "") != f"Bearer {secret}":
                return self.json_response({"ok": False, "error": "Unauthorized cron request"}, 401)
            try:
                return self.json_response(run_configured_discovery())
            except Exception as exc:
                return self.json_response({"ok": False, "error": str(exc)}, 400)
        match = re.fullmatch(r"/p/([A-Za-z0-9_-]{20,80})\.pdf", path)
        if match:
            cleanup_pdf_links()
            item = PDF_LINKS.get(match.group(1))
            if not item:
                return self.json_response({"ok": False, "error": "PDF link expired or not found"}, 404)
            item["downloads"] = int(item.get("downloads", 0)) + 1
            pdf = item["pdf"]
            self.send_response(200)
            self.send_header("Content-Type", "application/pdf")
            self.send_header("Content-Disposition", f'inline; filename="{item["filename"]}"')
            self.send_header("Content-Length", str(len(pdf)))
            self.send_header("X-Robots-Tag", "noindex, nofollow, noarchive")
            self.end_headers()
            self.wfile.write(pdf)
            return
        return super().do_GET()

    def do_POST(self):
        path = urlparse(self.path).path
        auth_error = integration_auth_error(self.headers, path)
        if auth_error:
            status, message = auth_error
            return self.json_response({"ok": False, "error": message}, status)
        if path not in {"/api/audit", "/api/ai-seo-review", "/api/analyze-clinic-candidates", "/api/send", "/api/vendor-search", "/api/clinic-search", "/api/import-search-html", "/api/enrich-clinics", "/api/scrape-directory", "/api/contact-enrich", "/api/exhibition/import", "/api/exhibition/enrich", "/api/leads/bulk", "/api/export-clinics", "/api/generate-article", "/api/proposal-pdf", "/api/proposal-link"}:
            return self.json_response({"ok": False, "error": "Not found"}, 404)
        try:
            length = int(self.headers.get("Content-Length", "0"))
            request_limit = 2_000_000 if path in {"/api/proposal-pdf", "/api/proposal-link", "/api/send", "/api/import-search-html", "/api/enrich-clinics", "/api/exhibition/import", "/api/exhibition/enrich", "/api/leads/bulk", "/api/export-clinics", "/api/analyze-clinic-candidates"} else 30_000
            if length <= 0 or length > request_limit:
                raise ValueError("Invalid request size")
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
            if path == "/api/export-clinics":
                content, filename, content_type = export_clinic_candidates(payload)
                self.send_response(200)
                self.send_header("Content-Type", content_type)
                self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
                self.send_header("Content-Length", str(len(content)))
                self.end_headers()
                self.wfile.write(content)
                return
            if path == "/api/proposal-link":
                token, filename, expires_at = store_proposal_link(payload)
                return self.json_response({"ok": True, "url": f"{self.public_base_url()}/p/{token}.pdf",
                                           "filename": filename, "expiresAt": expires_at,
                                           "ttlSeconds": PDF_LINK_TTL,
                                           "warning": "Temporary link; it expires and may be lost if a free container restarts."})
            if path == "/api/proposal-pdf":
                pdf, filename = make_proposal_pdf(payload)
                self.send_response(200)
                self.send_header("Content-Type", "application/pdf")
                self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
                self.send_header("Content-Length", str(len(pdf)))
                self.end_headers()
                self.wfile.write(pdf)
                return
            if path == "/api/ai-seo-review":
                return self.json_response(generate_ai_seo_review(payload))
            if path == "/api/analyze-clinic-candidates":
                return self.json_response(analyze_clinic_candidates_ai(payload))
            if path == "/api/send":
                return self.json_response(send_message(payload))
            if path == "/api/vendor-search":
                return self.json_response(search_vendors(payload))
            if path == "/api/clinic-search":
                return self.json_response(search_clinics(payload))
            if path == "/api/import-search-html":
                return self.json_response(parse_search_html(payload))
            if path == "/api/enrich-clinics":
                return self.json_response(enrich_clinic_candidates(payload))
            if path == "/api/scrape-directory":
                return self.json_response(scrape_clinic_directory(payload))
            if path == "/api/contact-enrich":
                return self.json_response(enrich_public_business_contacts(
                    str(payload.get("url", "")).strip(), int(payload.get("maxPages", 3) or 3)))
            if path == "/api/exhibition/import":
                return self.json_response(parse_exhibition_data(payload))
            if path == "/api/exhibition/enrich":
                return self.json_response(enrich_exhibition_companies(payload))
            if path == "/api/leads/bulk":
                items = payload.get("items") if isinstance(payload.get("items"), list) else []
                return self.json_response(persist_leads_database(items))
            if path == "/api/generate-article":
                return self.json_response(generate_seo_article(payload))
            url = str(payload.get("url", "")).strip()
            if not url:
                raise ValueError("URL is required")
            return self.json_response(audit(url))
        except (ValueError, URLError, HTTPError, socket.timeout, TimeoutError) as exc:
            return self.json_response({"ok": False, "error": str(exc), "type": type(exc).__name__}, 400)
        except Exception as exc:
            label = ("AI SEO review" if path == "/api/ai-seo-review" else
                     "AI clinic classification" if path == "/api/analyze-clinic-candidates" else
                     "Clinic export" if path == "/api/export-clinics" else
                     "Send" if path == "/api/send" else
                     "Vendor search" if path == "/api/vendor-search" else
                     "Clinic search" if path == "/api/clinic-search" else
                     "Search HTML import" if path == "/api/import-search-html" else
                     "Clinic enrichment" if path == "/api/enrich-clinics" else
                     "Directory scraper" if path == "/api/scrape-directory" else
                     "Public contact enrichment" if path == "/api/contact-enrich" else
                     "Exhibition import" if path == "/api/exhibition/import" else
                     "Exhibition enrichment" if path == "/api/exhibition/enrich" else
                     "Lead database" if path == "/api/leads/bulk" else
                     "Article generation" if path == "/api/generate-article" else
                     "Proposal PDF" if path in {"/api/proposal-pdf", "/api/proposal-link"} else "Audit")
            return self.json_response({"ok": False, "error": f"{label} failed: {type(exc).__name__}: {exc}"}, 500)


def main():
    host = os.getenv("HOST", "127.0.0.1")
    port = int(os.getenv("PORT", "8000"))
    server = ThreadingHTTPServer((host, port), Handler)
    print(f"Clinic Signal running at http://{host}:{port}")
    print("Press Ctrl+C to stop.")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
